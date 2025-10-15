from __future__ import annotations  # ← Debe ser la PRIMERA línea del archivo

# Autenticación y permisos
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from urllib.parse import urlencode
# HTTP y respuestas
from django.http import HttpRequest, HttpResponse, HttpResponseBadRequest, HttpResponseForbidden, JsonResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt

# Utilidades Django
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.timezone import now
from django.utils.encoding import force_str
from django.conf import settings

# Templates
from django.template.loader import render_to_string

# WeasyPrint compatible (nube OK, local Windows no rompe el server)
from core.utils.weasy_compat import HTML, CSS, WEASY_AVAILABLE, WEASY_IMPORT_ERROR
from core.utils.decorators import require_weasy  # si usas el decorador
from django.urls import reverse 


# Base de datos
from django.db import transaction, IntegrityError, connection
from core.db_errors import map_db_error   # <-- tu helper de validación de errores DB

# Extras para boletines y validaciones
from datetime import datetime, date
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from io import BytesIO
from django.db.models import Count, Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json
import re
from django.shortcuts import render
from django.views.decorators.cache import never_cache  # <-- importa esto

# imports de modelos (deben estar juntos)
from .models import Sede, Grupo, Estudiante, EstudianteGrupo
from django.db import connection
from django.utils import timezone
from django.urls import reverse
from django.db import connection, transaction
from django.contrib import messages
from django.views.decorators.http import require_http_methods

# =========================================================
# Login / Logout
# =========================================================
@never_cache
def login_view(request: HttpRequest) -> HttpResponse:
    """GET: login • POST: autentica y redirige por rol (respeta ?next=)."""
    if request.user.is_authenticated:
        return _redir_por_rol(request.user)

    error = None
    next_url = request.GET.get("next") or request.POST.get("next")

    if request.method == "POST":
        usuario = (request.POST.get("usuario") or "").strip()
        password = request.POST.get("password") or ""
        user = authenticate(request, username=usuario, password=password)
        if user is None:
            error = "Usuario o contraseña incorrectos"
        else:
            login(request, user)
            if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                return redirect(next_url)
            return _redir_por_rol(user)

    return render(request, "core/login.html", {"error": error, "next": next_url})

def logout_view(request: HttpRequest) -> HttpResponse:
    logout(request)
    return redirect("login")

def _redir_por_rol(user) -> HttpResponse:
    rol = (getattr(user, "rol", "") or "").upper()
    if rol == "RECTOR":
        return redirect("dashboard_rector")
    if rol == "DOCENTE":
        return redirect("dashboard_docente")
    return redirect("dashboard_admin")

# =========================================================
# Dashboards
# =========================================================
@login_required(login_url="login")
def dashboard_rector(request: HttpRequest) -> HttpResponse:
    return render(request, "core/dashboard_rector.html")


@login_required(login_url="login")
def dashboard_docente(request: HttpRequest) -> HttpResponse:
    return render(request, "core/dashboard_docente.html")


@login_required(login_url="login")
def dashboard_admin(request: HttpRequest) -> HttpResponse:
    # Ya existía, lo dejamos igual
    return render(request, "core/dashboard_admin.html")

@login_required(login_url="login")
def administrativo_reportes_academicos_filtro(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_admin(request)) is not None:
        return resp
    # Plantilla que acordamos:
    return render(request, "core/administrativo/administrativo_reportes_academicos_filtro.html")

# Resultado POR GRUPO (página que lee los parámetros por querystring y consume APIs)
@login_required(login_url="login")
def administrativo_reportes_academicos_por_grupo(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_admin(request)) is not None:
        return resp
    # Esta es la página “por grupo” (UI con exportar/tabla)
    return render(request, "core/administrativo/administrativo_reportes_academicos_por_grupo.html")

# (Opcional) Variante POR ESTUDIANTE (de momento conserva flujo por grupo)
@login_required(login_url="login")
def administrativo_reportes_academicos_por_estudiante(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_admin(request)) is not None:
        return resp
    return render(request, "core/administrativo/reportes_academicos_estudiante_admin.html")

# (Opcional) Tabla simple — si la usas como pantalla separada
@login_required(login_url="login")
def administrativo_reportes_academicos_tabla(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_admin(request)) is not None:
        return resp
    return render(request, "core/administrativo/reportes_academicos_tabla_admin.html")

def estilizar_boletin_ws(ws, ancho_columna_1=48, ancho_otras=18):
    """
    Aplica estilos "colegio": encabezado oscuro, cuerpo con bandas, bordes,
    formatos de números para notas y congelar encabezado.
    No altera los datos.
    """
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # Colores institucionales (coinciden con dashboard.css)
    HEADER_BG = "0F172A"  # navy oscuro
    HEADER_FG = "FFFFFF"
    BAND_BG   = "F8FAFC"  # fila alterna suave
    BORDER_CL = "D1D5DB"

    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_font = Font(bold=True, color=HEADER_FG, size=11)
    center      = Alignment(vertical="center", horizontal="center", wrap_text=True)
    left        = Alignment(vertical="center", horizontal="left", wrap_text=True)
    thin_side   = Side(border_style="thin", color=BORDER_CL)
    border      = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    # Filtro y congelar encabezado
    try:
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass
    ws.freeze_panes = ws["A2"]

    # Encabezado (fila 1)
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        # Anchos
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = ancho_columna_1 if c == 1 else ancho_otras

    # Cuerpo
    band_fill = PatternFill("solid", fgColor=BAND_BG)
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            # Bandas
            if r % 2 == 0:
                cell.fill = band_fill
            # Alineaciones y formato de nota (desde la 3a columna suelen ir las notas)
            if c == 1:
                cell.alignment = left
            else:
                cell.alignment = center
                # Si es número, le damos formato 0.00 (nota con 2 decimales)
                if isinstance(cell.value, (int, float)) and c >= 3:
                    cell.number_format = "0.00"

# =========================================================
# Helpers (acceso rector)
# =========================================================
def _solo_rector(request: HttpRequest) -> bool:
    return (getattr(request.user, "rol", "") or "").upper() == "RECTOR"


def _guard_rector(request: HttpRequest) -> HttpResponse | None:
    """Devuelve redirección si NO es rector; si es rector devuelve None."""
    if not _solo_rector(request):
        return _redir_por_rol(request.user)
    return None


# ========= Helpers de rol DOCENTE =========
def _solo_docente(request) -> bool:
    """True si el usuario autenticado tiene rol DOCENTE."""
    rol = (getattr(request.user, "rol", "") or "").upper()
    return rol == "DOCENTE"

def _guard_docente(request: HttpRequest) -> HttpResponse | None:
    """Devuelve redirección si NO es docente; si es docente devuelve None."""
    if not _solo_docente(request):
        return _redir_por_rol(request.user)
    return None

def _docente_puede_ver_grupo(usuario_id: int, grupo_id: int) -> bool:
    """Valida si el usuario (docente) está vinculado al grupo (docente_grupo)."""
    with connection.cursor() as cur:
        cur.execute("""
            SELECT 1
              FROM public.docentes d
              JOIN public.docente_grupo dg ON dg.docente_id = d.id
             WHERE d.usuario_id = %s AND dg.grupo_id = %s
             LIMIT 1;
        """, [usuario_id, grupo_id])
        return cur.fetchone() is not None

def _docente_puede_editar_asignatura(usuario_id: int, grupo_id: int, asignatura_id: int) -> bool:
    """Valida si el docente tiene esa asignatura en ese grupo (docente_asignacion)."""
    from django.db import connection
    with connection.cursor() as cur:
        cur.execute("""
            SELECT 1
              FROM public.docentes d
              JOIN public.docente_asignacion da ON da.docente_id = d.id
              JOIN public.grupo_asignatura ga ON ga.id = da.grupo_asignatura_id
             WHERE d.usuario_id = %s
               AND ga.grupo_id = %s
               AND ga.asignatura_id = %s
             LIMIT 1;
        """, [usuario_id, grupo_id, asignatura_id])
        return cur.fetchone() is not None

_ID = re.compile(r"^\d{1,10}$")
def _ok(x: str) -> bool: return bool(_ID.fullmatch((x or "").strip()))

@login_required(login_url="login")
@require_GET
def api_docente_sedes(request: HttpRequest) -> JsonResponse:
    if (resp := _guard_docente(request)) is not None:
        return JsonResponse({"sedes": []})
    uid = request.user.id
    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT s.id, s.nombre
              FROM public.docentes d
              JOIN public.docente_grupo dg ON dg.docente_id = d.id
              JOIN public.grupos g ON g.id = dg.grupo_id
              JOIN public.sedes s ON s.id = g.sede_id
             WHERE d.usuario_id = %s
             ORDER BY s.nombre;
        """, [uid])
        sedes = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"sedes": sedes})

@login_required(login_url="login")
def exportar_boletines(request, grupo_id):
    # Similar a lo de rector
    pass

@login_required(login_url="login")
def api_docente_grados_por_sede(request):
    if (resp := _guard_docente(request)) is not None:
        return JsonResponse({"grados": []})
    uid = request.user.id
    sede_id = (request.GET.get("sede_id") or "").strip()
    if not _ok(sede_id):
        return JsonResponse({"grados": []})
    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT gr.id, gr.nombre
              FROM public.docentes d
              JOIN public.docente_grupo dg ON dg.docente_id = d.id
              JOIN public.grupos g ON g.id = dg.grupo_id
              JOIN public.grados gr ON gr.id = g.grado_id
             WHERE d.usuario_id = %s AND g.sede_id = %s
             ORDER BY gr.nombre;
        """, [uid, sede_id])
        grados = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"grados": grados})

@login_required(login_url="login")
def api_docente_grupos_por_sede_grado(request):
    if (resp := _guard_docente(request)) is not None:
        return JsonResponse({"grupos": []})
    uid = request.user.id
    sede_id = (request.GET.get("sede_id") or "").strip()
    grado_id = (request.GET.get("grado_id") or "").strip()
    if not (_ok(sede_id) and _ok(grado_id)):
        return JsonResponse({"grupos": []})
    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT g.id, g.nombre
              FROM public.docentes d
              JOIN public.docente_grupo dg ON dg.docente_id = d.id
              JOIN public.grupos g ON g.id = dg.grupo_id
             WHERE d.usuario_id = %s
               AND g.sede_id = %s
               AND g.grado_id = %s
             ORDER BY g.nombre;
        """, [uid, sede_id, grado_id])
        grupos = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"grupos": grupos})

@login_required(login_url="login")
def api_docente_areas_por_grupo(request):
    if (resp := _guard_docente(request)) is not None:
        return JsonResponse({"areas": []})
    uid = request.user.id
    grupo_id = (request.GET.get("grupo_id") or "").strip()
    if not _ok(grupo_id):
        return JsonResponse({"areas": []})
    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT ar.id, ar.nombre
              FROM public.docentes d
              JOIN public.docente_asignacion da ON da.docente_id = d.id
              JOIN public.grupo_asignatura ga   ON ga.id = da.grupo_asignatura_id
              JOIN public.asignaturas asig      ON asig.id = ga.asignatura_id
              JOIN public.areas ar              ON ar.id = asig.area_id
             WHERE d.usuario_id = %s
               AND ga.grupo_id = %s
             ORDER BY ar.nombre;
        """, [uid, grupo_id])
        areas = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"areas": areas})

@login_required(login_url="login")
def api_docente_asignaturas_por_grupo_area(request):
    if (resp := _guard_docente(request)) is not None:
        return JsonResponse({"asignaturas": []})
    uid = request.user.id
    grupo_id = (request.GET.get("grupo_id") or "").strip()
    area_id  = (request.GET.get("area_id")  or "").strip()
    if not (_ok(grupo_id) and _ok(area_id)):
        return JsonResponse({"asignaturas": []})
    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT asig.id, asig.nombre
              FROM public.docentes d
              JOIN public.docente_asignacion da ON da.docente_id = d.id
              JOIN public.grupo_asignatura ga   ON ga.id = da.grupo_asignatura_id
              JOIN public.asignaturas asig      ON asig.id = ga.asignatura_id
             WHERE d.usuario_id = %s
               AND ga.grupo_id = %s
               AND asig.area_id = %s
             ORDER BY asig.nombre;
        """, [uid, grupo_id, area_id])
        asignaturas = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"asignaturas": asignaturas})

@login_required(login_url="login")
def api_docente_periodos_abiertos(request):
    if (resp := _guard_docente(request)) is not None:
        return JsonResponse({"periodos": []})
    with connection.cursor() as cur:
        cur.execute("""
            SELECT id, nombre
              FROM public.periodos
             WHERE abierto IS TRUE
             ORDER BY fecha_inicio;
        """)
        periodos = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"periodos": periodos})

@login_required(login_url="login")
@require_GET
def api_docente_estudiantes_por_grupo(request: HttpRequest) -> JsonResponse:
    if (resp := _guard_docente(request)) is not None:
        return resp
    uid = request.user.id
    grupo_id = (request.GET.get("grupo_id") or "").strip()
    if not _ok(grupo_id):
        return JsonResponse({"estudiantes": []})

    with connection.cursor() as cur:
        # Validar permiso al grupo
        cur.execute("""
            SELECT 1
            FROM public.docentes d
            JOIN public.docente_grupo dg ON dg.docente_id = d.id
            WHERE d.usuario_id = %s AND dg.grupo_id = %s
            LIMIT 1;
        """, [uid, grupo_id])
        if cur.fetchone() is None:
            return JsonResponse({"estudiantes": []})

        # Estudiantes activos en el grupo (fecha_fin IS NULL)
        cur.execute("""
            SELECT e.id, e.documento, e.nombre, e.apellidos
            FROM public.estudiantes e
            JOIN public.estudiante_grupo eg ON eg.estudiante_id = e.id
            WHERE eg.grupo_id = %s
            AND (eg.fecha_fin IS NULL)
            ORDER BY e.apellidos, e.nombre;
        """, [grupo_id])
        rows = cur.fetchall()

    estudiantes = [{"id": r[0], "documento": r[1], "nombre": r[2], "apellidos": r[3]} for r in rows]
    return JsonResponse({"estudiantes": estudiantes})

@login_required(login_url="login")
@require_GET
def api_docente_notas_por_grupo_asignatura_periodo(request):
    if (resp := _guard_docente(request)) is not None:
        return resp
    uid = request.user.id

    grupo_id      = (request.GET.get("grupo_id") or "").strip()
    asignatura_id = (request.GET.get("asignatura_id") or "").strip()
    periodo_id    = (request.GET.get("periodo_id") or "").strip()
    if not (_ok(grupo_id) and _ok(asignatura_id) and _ok(periodo_id)):
        return JsonResponse({"notas": []})

    with connection.cursor() as cur:
        # Validar que el docente dicta esa asignatura en ese grupo
        cur.execute("""
            SELECT 1
              FROM public.docentes d
              JOIN public.docente_asignacion da ON da.docente_id = d.id
              JOIN public.grupo_asignatura ga ON ga.id = da.grupo_asignatura_id
             WHERE d.usuario_id = %s
               AND ga.grupo_id = %s
               AND ga.asignatura_id = %s
             LIMIT 1;
        """, [uid, grupo_id, asignatura_id])
        if cur.fetchone() is None:
            return JsonResponse({"notas": []})

        # Notas de los estudiantes activos del grupo
        cur.execute("""
            SELECT e.id AS estudiante_id, n.nota, n.fallas
              FROM public.estudiantes e
              JOIN public.estudiante_grupo eg ON eg.estudiante_id = e.id
              LEFT JOIN public.notas n
                     ON n.estudiante_id = e.id
                    AND n.asignatura_id = %s
                    AND n.periodo_id    = %s
             WHERE eg.grupo_id = %s
               AND (eg.fecha_fin IS NULL)
             ORDER BY e.apellidos, e.nombre;
        """, [asignatura_id, periodo_id, grupo_id])
        rows = cur.fetchall()

    notas = [{"estudiante_id": r[0], "nota": r[1], "fallas": r[2]} for r in rows]
    return JsonResponse({"notas": notas})

# ===== DOCENTE: páginas de registro de notas =====
@login_required(login_url="login")
def docente_registro_notas_filtro(request):
    if (resp := _guard_docente(request)) is not None:
        return resp
    return render(request, "core/docente/registro_notas_filtro.html")


@login_required(login_url="login")
def docente_registro_notas_por_grupo(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_docente(request)) is not None:
        return resp
    return render(request, "core/docente/registro_notas_por_grupo.html")


# ===== DOCENTE: páginas de reportes =====
@login_required(login_url="login")
def docente_reportes_academicos_filtro(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_docente(request)) is not None:
        return resp
    return render(request, "core/docente/reportes_academicos_filtro.html")

@login_required(login_url="login")
def docente_reportes_academicos_por_grupo(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_docente(request)) is not None:
        return resp
    return render(request, "core/docente/reportes_academicos_por_grupo.html")


@login_required(login_url="login")
def docente_registrar_nota(request: HttpRequest) -> JsonResponse:
    # Solo docente
    if (resp := _guard_docente(request)) is not None:
        return JsonResponse({"ok": False, "msg": "No autorizado."}, status=403)

    # --------- INPUTS ---------
    estudiante_id = (request.POST.get("estudiante_id") or "").strip()
    asignatura_id = (request.POST.get("asignatura_id") or "").strip()
    periodo_id    = (request.POST.get("periodo_id") or "").strip()
    nota_str      = (request.POST.get("nota") or "").strip()
    fallas_str    = (request.POST.get("fallas") or "0").strip()

    # ids numéricos
    if not all(re.fullmatch(r"\d{1,10}", x or "") for x in (estudiante_id, asignatura_id, periodo_id)):
        return JsonResponse({"ok": False, "msg": "Parámetros inválidos."}, status=400)

    # fallas entero >= 0
    try:
        fallas = int(fallas_str)
        if fallas < 0:
            raise ValueError()
    except ValueError:
        return JsonResponse({"ok": False, "msg": "Las fallas deben ser un número entero ≥ 0."}, status=400)

    # nota Decimal 1.00..5.00
    try:
        nota = Decimal(nota_str).quantize(Decimal("0.01"))
        if nota < Decimal("1.00") or nota > Decimal("5.00"):
            raise InvalidOperation()
    except (InvalidOperation, ValueError):
        return JsonResponse({"ok": False, "msg": "La nota debe estar entre 1.00 y 5.00 con dos decimales."}, status=400)

    user_id = request.user.id
    fuente_rol = "DOCENTE"

    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                # Periodo ABIERTO
                cur.execute("SELECT 1 FROM public.periodos WHERE id=%s AND abierto IS TRUE;", [periodo_id])
                if cur.fetchone() is None:
                    return JsonResponse({"ok": False, "msg": "No puede registrar nota: el periodo no está abierto."}, status=400)

                # Grupo ACTIVO del estudiante
                cur.execute("""
                    SELECT eg.grupo_id
                      FROM public.estudiante_grupo eg
                     WHERE eg.estudiante_id=%s AND eg.fecha_fin IS NULL
                     LIMIT 1;
                """, [estudiante_id])
                row_grp = cur.fetchone()
                if row_grp is None:
                    return JsonResponse({"ok": False, "msg": "El estudiante no tiene un grupo activo."}, status=400)
                grupo_id = int(row_grp[0])

                # Asignatura ∈ grupo activo
                cur.execute("""
                    SELECT 1
                      FROM public.grupo_asignatura ga
                     WHERE ga.grupo_id=%s AND ga.asignatura_id=%s
                     LIMIT 1;
                """, [grupo_id, asignatura_id])
                if cur.fetchone() is None:
                    return JsonResponse({"ok": False, "msg": "La asignatura no pertenece al grupo activo del estudiante."}, status=400)

            # ====== Validación clave: el docente tiene esa asignatura en ese grupo ======
            if not _docente_puede_editar_asignatura(user_id, grupo_id, int(asignatura_id)):
                return JsonResponse({"ok": False, "msg": "No autorizado para registrar notas en esa asignatura/grupo."}, status=403)

            # ==== UPSERT en public.notas (misma lógica del rector) ====
            with connection.cursor() as cur:
                # Nota anterior (si existe)
                cur.execute("""
                    SELECT id, nota, fallas
                      FROM public.notas
                     WHERE estudiante_id=%s AND asignatura_id=%s AND periodo_id=%s
                     LIMIT 1;
                """, [estudiante_id, asignatura_id, periodo_id])
                row = cur.fetchone()

                if row is None:
                    # INSERT
                    cur.execute("""
                        INSERT INTO public.notas (estudiante_id, asignatura_id, periodo_id,
                                                  nota, fallas, actualizado_por_usuario, fuente_rol, actualizado_en)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, now())
                        RETURNING id;
                    """, [estudiante_id, asignatura_id, periodo_id, str(nota), fallas, user_id, fuente_rol])
                    nota_id = cur.fetchone()[0]
                    accion = "INSERT"

                    cur.execute("""
                        INSERT INTO public.notas_historial
                            (nota_id, estudiante_id, asignatura_id, periodo_id,
                             nota_anterior, fallas_anterior,
                             nota_nueva, fallas_nuevas,
                             accion, realizado_por_usuario, realizado_en)
                        VALUES (%s,%s,%s,%s, NULL, NULL, %s, %s, 'INSERT', %s, now());
                    """, [nota_id, estudiante_id, asignatura_id, periodo_id, str(nota), fallas, user_id])

                else:
                    # UPDATE
                    nota_id_ant, nota_ant, fallas_ant = row
                    cur.execute("""
                        UPDATE public.notas
                           SET nota=%s, fallas=%s,
                               actualizado_por_usuario=%s, fuente_rol=%s, actualizado_en=now()
                         WHERE id=%s
                         RETURNING id;
                    """, [str(nota), fallas, user_id, fuente_rol, nota_id_ant])
                    nota_id = cur.fetchone()[0]
                    accion = "UPDATE"

                    cur.execute("""
                        INSERT INTO public.notas_historial
                            (nota_id, estudiante_id, asignatura_id, periodo_id,
                             nota_anterior, fallas_anterior,
                             nota_nueva, fallas_nuevas,
                             accion, realizado_por_usuario, realizado_en)
                        VALUES (%s,%s,%s,%s, %s, %s, %s, %s, 'UPDATE', %s, now());
                    """, [nota_id, estudiante_id, asignatura_id, periodo_id,
                          nota_ant, fallas_ant, str(nota), fallas, user_id])

        return JsonResponse({"ok": True, "accion": accion, "nota_id": nota_id})
    except Exception as e:
        return JsonResponse({"ok": False, "msg": f"Error al guardar: {e}"},
                            status=500)
    

@login_required(login_url="login")
def docente_reportes_academicos_export(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_docente(request)) is not None:
        return HttpResponse("No autorizado.", status=403)

    formato       = (request.GET.get('formato') or 'pdf').lower()
    grupo_id      = (request.GET.get('grupo_id') or '').strip()
    periodo_id    = (request.GET.get('periodo_id') or '').strip()
    estudiante_id = (request.GET.get('estudiante_id') or '').strip() or None

    if not (grupo_id and periodo_id):
        return HttpResponse("Faltan parámetros obligatorios (grupo_id, periodo_id).", status=400)
    if formato not in ("pdf", "excel"):
        return HttpResponse("Formato inválido.", status=400)

    # Seguridad: validar que el docente esté vinculado al grupo
    try:
        gid = int(grupo_id)
    except ValueError:
        return HttpResponse("grupo_id inválido.", status=400)

    if not _docente_puede_ver_grupo(request.user.id, gid):
        return HttpResponse("No autorizado para exportar ese grupo.", status=403)

    # Reutilizamos TU builder intacto (misma salida que rector)
    boletines = build_boletines(
        grupo_id=grupo_id,
        grado_id=None,
        sede_id=None,
        periodo_id=periodo_id,
        estudiante_id=estudiante_id,
    )

    if formato == "excel":
        return exportar_boletines_excel(boletines)   # misma función auxiliar que ya usas
    else:
        return exportar_boletines_pdf(boletines)     # idem


# =========================================================
# Estudiantes a grupos (página + endpoints)
# =========================================================
@login_required(login_url="login")
def rector_estudiantes_a_grupos(request: HttpRequest) -> HttpResponse:
    """
    Muestra el formulario y entrega 'sedes' y 'grados' desde BD.
    (El front luego filtra grados por sede con /api/grados-por-sede/).
    """
    if (resp := _guard_rector(request)) is not None:
        return resp

    with connection.cursor() as cur:
        cur.execute("SELECT id, nombre FROM public.sedes ORDER BY nombre;")
        sedes = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
        cur.execute("SELECT id, nombre FROM public.grados ORDER BY id;")
        grados = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]

    return render(
        request,
        "core/rector/estudiantes_a_grupos.html",
        {"sedes": sedes, "grados": grados},
    )


@login_required(login_url="login")
@require_POST
def rector_estudiantes_a_grupos_asignar(request: HttpRequest) -> JsonResponse:
    """
    Asigna un estudiante (por documento) a un grupo.
    - Verifica que el grupo pertenezca a la sede/grado elegidos.
    - Si ya está activo en ese grupo, devuelve OK sin duplicar.
    - Si estaba en otro grupo, cierra la relación activa y crea una nueva.
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"ok": False, "msg": "No autorizado."}, status=403)

    # JSON o form-data
    if request.content_type and "application/json" in request.content_type:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except Exception:
            payload = {}
    else:
        payload = request.POST

    doc = (payload.get("documento") or "").strip()
    sede_id = (payload.get("sede_id") or payload.get("sede") or "").strip()
    grado_id = (payload.get("grado_id") or payload.get("grado") or "").strip()
    grupo_id = (payload.get("grupo_id") or payload.get("grupo") or "").strip()

    # Validaciones
    if not re.fullmatch(r"\d{4,30}", doc or ""):
        return JsonResponse({"ok": False, "msg": "Documento inválido."}, status=400)
    for key, val in (("sede", sede_id), ("grado", grado_id), ("grupo", grupo_id)):
        if not re.fullmatch(r"\d{1,10}", val or ""):
            return JsonResponse({"ok": False, "msg": f"{key.capitalize()} inválido."}, status=400)

    with transaction.atomic():
        with connection.cursor() as cur:
            # Estudiante
            cur.execute(
                "SELECT id, nombre, apellidos FROM public.estudiantes WHERE documento=%s;",
                [doc],
            )
            est = cur.fetchone()
            if not est:
                return JsonResponse(
                    {"ok": False, "msg": "El documento no existe en 'estudiantes'."},
                    status=404,
                )
            estudiante_id = est[0]

            # Grupo válido para sede+grado
            cur.execute(
                """
                SELECT id FROM public.grupos
                WHERE id=%s AND sede_id=%s AND grado_id=%s;
                """,
                [grupo_id, sede_id, grado_id],
            )
            if cur.fetchone() is None:
                return JsonResponse(
                    {"ok": False, "msg": "Grupo no coincide con la sede/grado seleccionados."},
                    status=400,
                )

            # ¿Ya está activo en ese mismo grupo?
            cur.execute(
                """
                SELECT id FROM public.estudiante_grupo
                WHERE estudiante_id=%s AND grupo_id=%s AND fecha_fin IS NULL;
                """,
                [estudiante_id, grupo_id],
            )
            ya = cur.fetchone()
            if ya:
                return JsonResponse(
                    {"ok": True, "msg": "El estudiante ya estaba asignado a ese grupo.", "id": ya[0]}
                )

            # Cerrar asignaciones activas (si las hay)
            cur.execute(
                "UPDATE public.estudiante_grupo SET fecha_fin=CURRENT_DATE WHERE estudiante_id=%s AND fecha_fin IS NULL;",
                [estudiante_id],
            )

            # Insertar nueva
            cur.execute(
                """
                INSERT INTO public.estudiante_grupo (estudiante_id, grupo_id)
                VALUES (%s, %s)
                RETURNING id;
                """,
                [estudiante_id, grupo_id],
            )
            nuevo_id = cur.fetchone()[0]

    return JsonResponse({"ok": True, "msg": "Estudiante asignado con éxito.", "id": nuevo_id})


# =========================================================
# Asignación de DOCENTES a grupos (página + POST)
# =========================================================
@login_required(login_url="login")
def rector_asignacion_docentes_grupos(request: HttpRequest) -> HttpResponse:
    """
    Carga:
      - Docentes (join docentes + usuarios) para mostrar nombre completo y usuario.
      - Grupos (join grupos + sedes + grados) para rotular sede/grado/grupo.
    """
    if (resp := _guard_rector(request)) is not None:
        return resp

    with connection.cursor() as cur:
        # Docentes: id de 'docentes' + nombre/apellidos/usuario desde 'usuarios'
        cur.execute(
            """
            SELECT d.id, u.nombre, u.apellidos, u.usuario
            FROM public.docentes d
            JOIN public.usuarios u ON u.id = d.usuario_id
            ORDER BY u.nombre, u.apellidos;
            """
        )
        docentes = [
            {"id": r[0], "nombre": r[1], "apellidos": r[2], "usuario": r[3]}
            for r in cur.fetchall()
        ]

        # Grupos con etiquetas legibles (sede, grado, nombre de grupo)
        cur.execute(
            """
            SELECT g.id, s.nombre AS sede, gr.nombre AS grado, g.nombre AS grupo
            FROM public.grupos g
            JOIN public.sedes  s  ON s.id  = g.sede_id
            JOIN public.grados gr ON gr.id = g.grado_id
            ORDER BY s.nombre, gr.id, g.nombre;
            """
        )
        grupos = [
            {"id": r[0], "sede": r[1], "grado": r[2], "nombre": r[3]}
            for r in cur.fetchall()
        ]

    return render(
        request,
        "core/rector/asignacion_docentes_grupos.html",
        {"docentes": docentes, "grupos": grupos},
    )


@login_required(login_url="login")
@require_POST
def rector_asignacion_docentes_grupos_asignar(request: HttpRequest) -> JsonResponse:
    """
    Asigna un docente a un grupo:
      - Verifica que el grupo sea de la misma sede del docente.
      - Inserta en public.docente_grupo (único por docente+grupo).
      - Si ya existe, lo reporta sin duplicar.
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"ok": False, "msg": "No autorizado."}, status=403)

    docente_id = (request.POST.get("docente_id") or "").strip()
    grupo_id   = (request.POST.get("grupo_id") or "").strip()

    if not re.fullmatch(r"\d{1,10}", docente_id or "") or not re.fullmatch(r"\d{1,10}", grupo_id or ""):
        return JsonResponse({"ok": False, "msg": "Parámetros inválidos."}, status=400)

    with transaction.atomic():
        with connection.cursor() as cur:
            # Sede del docente
            cur.execute(
                """
                SELECT u.sede_id
                FROM public.docentes d
                JOIN public.usuarios u ON u.id = d.usuario_id
                WHERE d.id = %s;
                """,
                [docente_id],
            )
            row = cur.fetchone()
            if not row or row[0] is None:
                return JsonResponse({"ok": False, "msg": "El docente no tiene sede asociada."}, status=400)
            sede_doc = row[0]

            # Grupo pertenece a esa sede
            cur.execute("SELECT sede_id FROM public.grupos WHERE id=%s;", [grupo_id])
            r2 = cur.fetchone()
            if not r2:
                return JsonResponse({"ok": False, "msg": "Grupo inexistente."}, status=404)
            if r2[0] != sede_doc:
                return JsonResponse({"ok": False, "msg": "El grupo no pertenece a la sede del docente."}, status=400)

            # Insertar (evita duplicados por UNIQUE)
            cur.execute(
                """
                INSERT INTO public.docente_grupo (docente_id, grupo_id)
                VALUES (%s, %s)
                ON CONFLICT (docente_id, grupo_id) DO NOTHING
                RETURNING id;
                """,
                [docente_id, grupo_id],
            )
            inserted = cur.fetchone()

    if inserted:
        return JsonResponse({"ok": True, "msg": "Docente asignado al grupo con éxito.", "id": inserted[0]})
    else:
        return JsonResponse({"ok": True, "msg": "El docente ya estaba asignado a ese grupo."})


# =========================================================
# Registro de estudiantes (página + creación)
# =========================================================
@login_required(login_url="login")
def rector_registro_estudiantes(request: HttpRequest) -> HttpResponse:
    """Página de registro (el botón 'Crear' llama al endpoint JSON)."""
    if (resp := _guard_rector(request)) is not None:
        return resp
    return render(request, "core/rector/registro_estudiantes.html")


@login_required(login_url="login")
@require_POST
def rector_registro_estudiantes_crear(request: HttpRequest) -> JsonResponse:
    """
    Crea un estudiante con (nombre, apellidos, documento) en public.estudiantes.
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse(
            {"ok": False, "msg": "Solo el rector puede registrar estudiantes."},
            status=403,
        )

    # JSON o form-data
    if request.content_type and "application/json" in request.content_type:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except Exception:
            payload = {}
    else:
        payload = request.POST

    nombre = (payload.get("nombre") or "").strip()
    apellidos = (payload.get("apellidos") or "").strip()
    documento = (payload.get("documento") or "").strip()

    if not nombre or not apellidos or not documento:
        return JsonResponse({"ok": False, "msg": "Todos los campos son obligatorios."}, status=400)

    # Solo letras (incluye tildes y espacios simples)
    solo_letras = re.compile(r"^[A-Za-zÁÉÍÓÚÑáéíóúñ ]{2,80}$")
    if not solo_letras.fullmatch(nombre) or not solo_letras.fullmatch(apellidos):
        return JsonResponse({"ok": False, "msg": "Nombre y Apellidos deben contener solo letras."}, status=400)

    if not re.fullmatch(r"\d{5,20}", documento):
        return JsonResponse({"ok": False, "msg": "El documento debe tener entre 5 y 20 dígitos."}, status=400)

    sql = """
        INSERT INTO public.estudiantes (documento, nombre, apellidos)
        VALUES (%s, %s, %s)
        ON CONFLICT (documento) DO NOTHING
        RETURNING id;
    """
    with transaction.atomic():
        with connection.cursor() as cur:
            cur.execute(sql, [documento, nombre, apellidos])
            row = cur.fetchone()

    if row is None:
        return JsonResponse({"ok": False, "msg": "El documento ya existe."}, status=409)

    return JsonResponse({"ok": True, "msg": "Estudiante registrado con éxito.", "id": row[0]})


# =========================================================
# Registro de notas (filtro + variantes + tabla)
# =========================================================
@login_required(login_url="login")
def rector_registro_notas_filtro(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_rector(request)) is not None:
        return resp
    return render(request, "core/rector/registro_notas_filtro.html")


@login_required(login_url="login")
def rector_registro_notas_por_grupo(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_rector(request)) is not None:
        return resp
    return render(request, "core/rector/registro_notas_por_grupo.html")


@login_required(login_url="login")
def rector_registro_notas_por_estudiante(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_rector(request)) is not None:
        return resp
    return render(request, "core/rector/registro_notas_por_estudiante.html")


@login_required(login_url="login")
def rector_reporte_notas_tabla(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_rector(request)) is not None:
        return resp
    filtros = _build_filtros_tag(request)
    return render(request, "core/rector/reporte_notas_tabla.html", {"filtros": filtros})


# =========================================================
# Reportes académicos (filtro + variantes + tabla)
# =========================================================
@login_required(login_url="login")
def rector_reportes_academicos_filtro(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_rector(request)) is not None:
        return resp
    return render(request, "core/rector/reportes_academicos_filtro.html")


@login_required(login_url="login")
def rector_reportes_academicos_por_grupo(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_rector(request)) is not None:
        return resp
    return render(request, "core/rector/reportes_academicos_por_grupo.html")


@login_required(login_url="login")
def rector_reportes_academicos_por_estudiante(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_rector(request)) is not None:
        return resp
    return render(request, "core/rector/reportes_academicos_por_estudiante.html")


@login_required(login_url="login")
def rector_reportes_academicos_tabla(request: HttpRequest) -> HttpResponse:
    if (resp := _guard_rector(request)) is not None:
        return resp
    filtros = _build_filtros_tag(request)
    return render(request, "core/rector/reportes_academicos_tabla.html", {"filtros": filtros})


# =========================================================
# APIs reales (solo rector)
# =========================================================
@login_required(login_url="login")
@require_GET
def api_estudiante_por_documento(request: HttpRequest) -> JsonResponse:
    """
    Busca estudiante por documento en public.estudiantes.
    Respuesta: {id, documento, nombre, apellidos, nombre_completo}
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)

    doc = (request.GET.get("doc") or "").strip()
    if not re.fullmatch(r"\d{4,30}", doc or ""):
        return JsonResponse({"detail": "Parámetro 'doc' inválido."}, status=400)

    with connection.cursor() as cur:
        cur.execute(
            "SELECT id, documento, nombre, apellidos FROM public.estudiantes WHERE documento=%s;",
            [doc],
        )
        r = cur.fetchone()

    if not r:
        return JsonResponse({"detail": "No encontrado"}, status=404)

    return JsonResponse(
        {
            "id": r[0],
            "documento": r[1],
            "nombre": r[2],
            "apellidos": r[3],
            "nombre_completo": f"{r[2]} {r[3]}".strip(),
        }
    )

# === NUEVO: validar que el estudiante (por documento) esté activo en el grupo seleccionado ===
@login_required(login_url="login")
@require_GET
def api_estudiante_en_grupo_por_documento(request: HttpRequest) -> JsonResponse:
    """
    Valida y devuelve al estudiante por documento SOLO si está activo en el grupo dado.
    Parámetros: ?grupo_id=<id>&doc=<documento>
    Respuesta 200 OK:
      {"ok": true, "estudiante": {"id", "documento", "nombre", "apellidos", "nombre_completo"}}
    404 si no existe o no pertenece al grupo activo.
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)

    grupo_id = (request.GET.get("grupo_id") or "").strip()
    doc = (request.GET.get("doc") or "").strip()

    if not re.fullmatch(r"\d{1,10}", grupo_id or ""):
        return JsonResponse({"detail": "Parámetro 'grupo_id' inválido."}, status=400)
    if not re.fullmatch(r"\d{4,30}", doc or ""):
        return JsonResponse({"detail": "Parámetro 'doc' inválido."}, status=400)

    with connection.cursor() as cur:
        # Buscar estudiante por documento
        cur.execute(
            "SELECT id, documento, nombre, apellidos FROM public.estudiantes WHERE documento=%s;",
            [doc],
        )
        est = cur.fetchone()
        if not est:
            return JsonResponse({"detail": "Estudiante no encontrado."}, status=404)
        est_id, documento, nombre, apellidos = est

        # Verificar que esté activo en ese grupo
        cur.execute(
            """
            SELECT 1
              FROM public.estudiante_grupo
             WHERE estudiante_id=%s AND grupo_id=%s AND fecha_fin IS NULL
            """,
            [est_id, grupo_id],
        )
        ok = cur.fetchone() is not None

    if not ok:
        return JsonResponse({"detail": "El estudiante no pertenece al grupo indicado."}, status=404)

    return JsonResponse({
        "ok": True,
        "estudiante": {
            "id": est_id,
            "documento": documento,
            "nombre": nombre,
            "apellidos": apellidos,
            "nombre_completo": f"{nombre} {apellidos}".strip(),
        },
    })


@login_required(login_url="login")
@require_GET
def api_grados_por_sede(request: HttpRequest) -> JsonResponse:
    """
    Retorna los grados disponibles para una sede (según existan grupos).
    Respuesta: {"ok": true, "grados": [{"id":..., "nombre":"..."}, ...]}
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)

    sede_id = (request.GET.get("sede_id") or "").strip()
    if not re.fullmatch(r"\d{1,10}", sede_id or ""):
        return JsonResponse({"detail": "Parámetro 'sede_id' inválido."}, status=400)

    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT DISTINCT gr.id, gr.nombre
            FROM public.grupos g
            JOIN public.grados gr ON gr.id = g.grado_id
            WHERE g.sede_id = %s
            ORDER BY gr.id;
            """,
            [sede_id],
        )
        grados = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]

    return JsonResponse({"ok": True, "grados": grados})


@login_required(login_url="login")
@require_GET
def api_grupos_por_sede_grado(request: HttpRequest) -> JsonResponse:
    """
    Lista grupos para sede_id + grado_id.
    Respuesta: {"ok": true, "grupos": [{"id":..., "nombre":"..."}, ...]}
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)

    sede_id = (request.GET.get("sede_id") or "").strip()
    grado_id = (request.GET.get("grado_id") or "").strip()

    if not re.fullmatch(r"\d{1,10}", sede_id or "") or not re.fullmatch(r"\d{1,10}", grado_id or ""):
        return JsonResponse({"detail": "Parámetros inválidos."}, status=400)

    with connection.cursor() as cur:
        cur.execute(
            "SELECT id, nombre FROM public.grupos WHERE sede_id=%s AND grado_id=%s ORDER BY nombre;",
            [sede_id, grado_id],
        )
        datos = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]

    return JsonResponse({"ok": True, "grupos": datos})


# =========================================================
# Utilidad simple para mostrar filtros en el encabezado
# =========================================================
def _build_filtros_tag(request: HttpRequest) -> str:
    partes: list[str] = []

    sede = request.GET.get("sede")
    grado = request.GET.get("grado")
    grupo = request.GET.get("grupo")
    area = request.GET.get("area")
    asignatura = request.GET.get("asignatura")
    periodo = request.GET.get("periodo")
    docente = request.GET.get("docente")
    estudiante = request.GET.get("estudiante")

    if sede: partes.append(f"Sede {sede}")
    if grado: partes.append(f"Grado {grado}")
    if grupo: partes.append(f"Grupo {grupo}")
    if area: partes.append(f"Área {area}")
    if asignatura: partes.append(f"Asignatura {asignatura}")
    if periodo: partes.append(f"Periodo {periodo}")
    if docente: partes.append(f"Docente {docente}")
    if estudiante: partes.append(f"Estudiante {estudiante}")

    return " • ".join(partes) if partes else "Sin filtros"


# ========= NUEVO: APIs y POST para Asignación de Docentes a grupos =========
@login_required(login_url="login")
@require_GET
def api_docentes(request: HttpRequest) -> JsonResponse:
    """
    Lista docentes con su nombre completo y sede (derivada de usuarios.sede_id).
    Devuelve: {"ok": true, "docentes": [{"id":..., "nombre_completo":"...", "sede_id":...}, ...]}
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)

    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT d.id, u.nombre, u.apellidos, u.sede_id
            FROM public.docentes d
            JOIN public.usuarios u ON u.id = d.usuario_id
            ORDER BY u.apellidos, u.nombre;
            """
        )
        rows = cur.fetchall()

    docentes = [
        {"id": r[0], "nombre_completo": f"{r[1]} {r[2]}".strip(), "sede_id": r[3]}
        for r in rows
    ]
    return JsonResponse({"ok": True, "docentes": docentes})


@login_required(login_url="login")
@require_GET
def api_grupos_por_docente(request: HttpRequest) -> JsonResponse:
    """
    Dado un docente_id, devuelve TODOS los grupos de la sede del docente.
    Respuesta:
      {
        "ok": true,
        "grupos": [
          {"id": 5, "etiqueta": "11° 102", "grado": "11°", "grupo": "102", "sede_id": 3}
        ]
      }
    """
    # Solo rector
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)

    docente_id = (request.GET.get("docente_id") or "").strip()
    if not re.fullmatch(r"\d{1,10}", docente_id):
        return JsonResponse({"detail": "Parámetro 'docente_id' inválido."}, status=400)

    with connection.cursor() as cur:
        # 1) Obtener la sede del docente (vía usuarios)
        cur.execute(
            """
            SELECT u.sede_id
            FROM public.docentes d
            JOIN public.usuarios u ON u.id = d.usuario_id
            WHERE d.id = %s
            """,
            [docente_id],
        )
        row = cur.fetchone()
        if not row or row[0] is None:
            # Sin sede, no hay grupos que listar
            return JsonResponse({"ok": True, "grupos": []})

        sede_id = row[0]

        # 2) Traer TODOS los grupos de esa sede (ordenados por grado y nombre de grupo)
        cur.execute(
            """
            SELECT g.id,
                   gr.nombre AS grado,
                   g.nombre  AS grupo
            FROM public.grupos g
            JOIN public.grados gr ON gr.id = g.grado_id
            WHERE g.sede_id = %s
            ORDER BY gr.id, g.nombre
            """,
            [sede_id],
        )
        rows = cur.fetchall()

    grupos = [
        {
            "id": r[0],
            "etiqueta": f"{(r[1] or '').strip()} {(r[2] or '').strip()}".strip(),  # p.ej. "11° 102"
            "grado": (r[1] or "").strip(),
            "grupo": (r[2] or "").strip(),
            "sede_id": sede_id,
        }
        for r in rows
    ]

    return JsonResponse({"ok": True, "grupos": grupos})


@login_required(login_url="login")
@require_POST
def rector_vincular_docente_grupo(request: HttpRequest) -> JsonResponse:
    # Sólo rector
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"ok": False, "msg": "No autorizado."}, status=403)

    docente_id = (request.POST.get("docente_id") or "").strip()
    grupo_id   = (request.POST.get("grupo_id") or "").strip()

    # ids numéricos
    if not re.fullmatch(r"\d{1,10}", docente_id) or not re.fullmatch(r"\d{1,10}", grupo_id):
        return JsonResponse({"ok": False, "msg": "Parámetros inválidos."}, status=400)

    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                # Crea el vínculo si no existe (respeta tu UNIQUE)
                cur.execute("""
                    INSERT INTO public.docente_grupo (docente_id, grupo_id)
                    VALUES (%s, %s)
                    ON CONFLICT (docente_id, grupo_id) DO NOTHING
                    RETURNING id;
                """, [docente_id, grupo_id])
                created = cur.fetchone() is not None

        return JsonResponse({
            "ok": True,
            "created": created,
            "msg": "Vínculo docente↔grupo creado." if created else "El vínculo ya existía."
        })
    except Exception as e:
        ok, msg, code, http = map_db_error(e)
        return JsonResponse({"ok": False, "msg": msg or "No se pudo vincular."}, status=http or 400)


# ===== API: asignaturas por grupo (devuelve id de ga y nombre legible) =====
@login_required(login_url="login")
@require_GET
def api_asignaturas_por_grupo(request: HttpRequest) -> JsonResponse:
    """
    Para selección de asignaturas en registro de notas.
    Devuelve:
      - id    → ID real de la asignatura (usar para registrar/consultar notas)
      - ga_id → ID de grupo_asignatura (útil para asignación de docentes)
      - nombre
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)

    grupo_id = (request.GET.get("grupo_id") or "").strip()
    if not re.fullmatch(r"\d{1,10}", grupo_id or ""):
        return JsonResponse({"detail": "Parámetro 'grupo_id' inválido."}, status=400)

    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT ga.id AS ga_id, a.id AS asignatura_id, a.nombre
            FROM public.grupo_asignatura ga
            JOIN public.asignaturas a ON a.id = ga.asignatura_id
            WHERE ga.grupo_id = %s
            ORDER BY a.nombre;
            """,
            [grupo_id],
        )
        data = [{"ga_id": r[0], "id": r[1], "nombre": r[2]} for r in cur.fetchall()]
    return JsonResponse({"ok": True, "asignaturas": data})

# ===== POST: Asignar docente ↔ asignatura (en un grupo) =====
@login_required(login_url="login")
@require_POST
def rector_asignar_docente_asignatura(request: HttpRequest) -> JsonResponse:
    """
    Asigna un docente a una (grupo, asignatura) a través de grupo_asignatura (ga_id).

    Params (form-urlencoded o JSON):
      - docente_id (int)
      - grupo_id (int)
      - grupo_asignatura_id (int)
      - reemplazar (opcional: "1" para sacar al docente anterior si ya había uno)

    Reglas:
      - ga_id debe pertenecer a grupo_id
      - el docente debe tener asignado ese grupo (docente_grupo)
      - si ya existe un docente para ese ga_id:
          - si es el mismo -> OK idempotente
          - si es otro:
              - si reemplazar != "1" -> 409 con info del docente actual
              - si reemplazar == "1" -> elimina al anterior e inserta el nuevo
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"ok": False, "msg": "No autorizado."}, status=403)

    # admitir JSON o form-data
    if request.content_type and "application/json" in (request.content_type or ""):
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except Exception:
            payload = {}
    else:
        payload = request.POST

    docente_id = (str(payload.get("docente_id") or "").strip())
    grupo_id   = (str(payload.get("grupo_id") or "").strip())
    ga_id      = (str(payload.get("grupo_asignatura_id") or payload.get("ga_id") or "").strip())
    reemplazar = (str(payload.get("reemplazar") or "").strip())

    # Validaciones básicas
    if not all(re.fullmatch(r"\d{1,10}", v or "") for v in [docente_id, grupo_id, ga_id]):
        return JsonResponse({"ok": False, "msg": "Parámetros inválidos."}, status=400)

    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                # 1) Confirmar que ga pertenece al grupo indicado
                cur.execute("SELECT grupo_id, asignatura_id FROM public.grupo_asignatura WHERE id=%s;", [ga_id])
                row = cur.fetchone()
                if not row:
                    return JsonResponse({"ok": False, "msg": "La asignatura no existe para ese grupo."}, status=404)
                if str(row[0]) != grupo_id:
                    return JsonResponse({"ok": False, "msg": "Inconsistencia: la asignatura no es de ese grupo."}, status=400)

                # 2) Confirmar que el docente puede ver ese grupo (docente_grupo)
                cur.execute(
                    "SELECT 1 FROM public.docente_grupo WHERE docente_id=%s AND grupo_id=%s;",
                    [docente_id, grupo_id],
                )
                if cur.fetchone() is None:
                    return JsonResponse({"ok": False, "msg": "El docente no tiene asignado ese grupo."}, status=400)

                # 3) ¿Ya hay un docente para este (grupo, asignatura)?
                cur.execute(
                    """
                    SELECT da.docente_id, u.nombre, u.apellidos
                    FROM public.docente_asignacion da
                    JOIN public.docentes d ON d.id = da.docente_id
                    JOIN public.usuarios u ON u.id = d.usuario_id
                    WHERE da.grupo_asignatura_id = %s
                    LIMIT 1;
                    """,
                    [ga_id],
                )
                ya = cur.fetchone()

                if ya:
                    docente_actual_id, nom, ape = ya
                    # Si es el mismo docente → idempotente
                    if str(docente_actual_id) == docente_id:
                        return JsonResponse({"ok": True, "msg": "El docente ya estaba asignado a esta asignatura del grupo."})

                    # Si es otro y NO pidió reemplazar → 409 con info del actual
                    if reemplazar != "1":
                        return JsonResponse(
                            {
                                "ok": False,
                                "msg": f"Ya existe un docente asignado ({nom} {ape}). "
                                       f"Para reemplazarlo envía reemplazar=1.",
                                "docente_actual": {
                                    "id": int(docente_actual_id),
                                    "nombre_completo": f"{nom} {ape}".strip(),
                                }
                            },
                            status=409,
                        )

                    # Reemplazar: borrar al actual y seguir
                    cur.execute(
                        "DELETE FROM public.docente_asignacion WHERE grupo_asignatura_id=%s;",
                        [ga_id],
                    )

                # 4) Insertar evitando duplicados (UNIQUE)
                cur.execute(
                    """
                    INSERT INTO public.docente_asignacion (docente_id, grupo_asignatura_id)
                    VALUES (%s, %s)
                    ON CONFLICT (docente_id, grupo_asignatura_id) DO NOTHING
                    RETURNING id;
                    """,
                    [docente_id, ga_id],
                )
                inserted = cur.fetchone()

        if inserted:
            return JsonResponse({"ok": True, "msg": "Asignatura asignada al docente con éxito.", "id": inserted[0]})
        return JsonResponse({"ok": True, "msg": "El docente ya tenía esa asignatura en el grupo."})

    except IntegrityError as e:
        ok, payload = map_db_error(e)
        if not ok:
            return JsonResponse(payload, status=payload.get("status", 400))
        # Si no fue mapeable, caemos al genérico:
        return JsonResponse({"ok": False, "msg": "Error de integridad."}, status=400)
    except Exception as e:
        return JsonResponse({"ok": False, "msg": f"Error inesperado: {e}"}, status=500)


# ===== POST: Quitar docente de una asignatura del grupo =====
@login_required(login_url="login")
@require_POST
def rector_quitar_docente_asignatura(request: HttpRequest) -> JsonResponse:
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"ok": False, "msg": "No autorizado."}, status=403)

    docente_id = (request.POST.get("docente_id") or "").strip()
    ga_id      = (request.POST.get("grupo_asignatura_id") or "").strip()
    if not re.fullmatch(r"\d{1,10}", docente_id or "") or not re.fullmatch(r"\d{1,10}", ga_id or ""):
        return JsonResponse({"ok": False, "msg": "Parámetros inválidos."}, status=400)

    with transaction.atomic():
        with connection.cursor() as cur:
            cur.execute(
                "DELETE FROM public.docente_asignacion WHERE docente_id=%s AND grupo_asignatura_id=%s RETURNING id;",
                [docente_id, ga_id],
            )
            deleted = cur.fetchone()

    if deleted:
        return JsonResponse({"ok": True, "msg": "Asignación eliminada."})
    return JsonResponse({"ok": True, "msg": "No había asignación para eliminar."})

# ====== NUEVO: APIs para filtros y notas por grupo ======

@login_required(login_url="login")
@require_GET
def api_sedes(request: HttpRequest) -> JsonResponse:
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)
    with connection.cursor() as cur:
        cur.execute("SELECT id, nombre FROM public.sedes ORDER BY nombre;")
        sedes = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"ok": True, "sedes": sedes})

@login_required(login_url="login")
@require_GET
def api_periodos_abiertos(request: HttpRequest) -> JsonResponse:
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)
    with connection.cursor() as cur:
        cur.execute("SELECT id, nombre FROM public.periodos WHERE abierto=true ORDER BY id;")
        periodos = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"ok": True, "periodos": periodos})

@login_required(login_url="login")
@require_GET
def api_areas_por_grupo(request: HttpRequest) -> JsonResponse:
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)
    grupo_id = (request.GET.get("grupo_id") or "").strip()
    if not re.fullmatch(r"\d{1,10}", grupo_id or ""):
        return JsonResponse({"detail": "Parámetro 'grupo_id' inválido."}, status=400)

    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT ar.id, ar.nombre
            FROM public.grupo_asignatura ga
            JOIN public.asignaturas a ON a.id = ga.asignatura_id
            JOIN public.areas ar ON ar.id = a.area_id
            WHERE ga.grupo_id = %s
            ORDER BY ar.nombre;
        """, [grupo_id])
        areas = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"ok": True, "areas": areas})

@login_required(login_url="login")
@require_GET
def api_asignaturas_por_grupo_area(request: HttpRequest) -> JsonResponse:
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)
    grupo_id = (request.GET.get("grupo_id") or "").strip()
    area_id  = (request.GET.get("area_id") or "").strip()
    if not re.fullmatch(r"\d{1,10}", grupo_id or "") or not re.fullmatch(r"\d{1,10}", area_id or ""):
        return JsonResponse({"detail": "Parámetros inválidos."}, status=400)

    with connection.cursor() as cur:
        cur.execute("""
            SELECT a.id, a.nombre
            FROM public.grupo_asignatura ga
            JOIN public.asignaturas a ON a.id = ga.asignatura_id
            WHERE ga.grupo_id = %s AND a.area_id = %s
            ORDER BY a.nombre;
        """, [grupo_id, area_id])
        asignaturas = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"ok": True, "asignaturas": asignaturas})

@login_required(login_url="login")
@require_GET
def api_estudiantes_por_grupo(request: HttpRequest) -> JsonResponse:
    """Lista estudiantes activos del grupo (orden: apellidos)."""
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)
    grupo_id = (request.GET.get("grupo_id") or "").strip()
    if not re.fullmatch(r"\d{1,10}", grupo_id or ""):
        return JsonResponse({"detail": "Parámetro 'grupo_id' inválido."}, status=400)

    with connection.cursor() as cur:
        cur.execute("""
            SELECT e.id, e.apellidos, e.nombre, e.documento
            FROM public.estudiante_grupo eg
            JOIN public.estudiantes e ON e.id = eg.estudiante_id
            WHERE eg.grupo_id=%s AND eg.fecha_fin IS NULL
            ORDER BY e.apellidos, e.nombre;
        """, [grupo_id])
        estudiantes = [{"id": r[0], "apellidos": r[1], "nombre": r[2], "documento": r[3]} for r in cur.fetchall()]
    return JsonResponse({"ok": True, "estudiantes": estudiantes})

@login_required(login_url="login")
@require_GET
def api_docente_de_grupo_asignatura(request: HttpRequest) -> JsonResponse:
    """
    Devuelve el docente (si existe) que dicta una asignatura en un grupo.
    Parámetros: grupo_id, asignatura_id
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)

    grupo_id = (request.GET.get("grupo_id") or "").strip()
    asignatura_id = (request.GET.get("asignatura_id") or "").strip()
    if not re.fullmatch(r"\d{1,10}", grupo_id or "") or not re.fullmatch(r"\d{1,10}", asignatura_id or ""):
        return JsonResponse({"detail": "Parámetros inválidos."}, status=400)

    with connection.cursor() as cur:
        cur.execute("""
            SELECT d.id, u.nombre, u.apellidos
            FROM public.docente_asignacion da
            JOIN public.grupo_asignatura ga ON ga.id = da.grupo_asignatura_id
            JOIN public.docentes d ON d.id = da.docente_id
            JOIN public.usuarios u ON u.id = d.usuario_id
            WHERE ga.grupo_id = %s AND ga.asignatura_id = %s
            LIMIT 1;
        """, [grupo_id, asignatura_id])
        r = cur.fetchone()

    if not r:
        return JsonResponse({"ok": True, "docente": None})

    return JsonResponse({"ok": True, "docente": {
        "id": r[0],
        "nombre": r[1],
        "apellidos": r[2],
    }})


# ====== NUEVO: obtener notas ya guardadas para precargar la tabla ======
@login_required(login_url="login")
@require_GET
def api_notas_por_grupo_asignatura_periodo(request: HttpRequest) -> JsonResponse:
    """
    Devuelve las notas/fallas ya guardadas para los estudiantes ACTIVOS del grupo,
    en una asignatura y periodo específicos.

    Parámetros: ?grupo_id=<id>&asignatura_id=<id>&periodo_id=<id>
    Respuesta:
      {"ok": true, "notas": [{"estudiante_id": 123, "nota": 4.3, "fallas": 2}, ...]}
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)

    grupo_id      = (request.GET.get("grupo_id") or "").strip()
    asignatura_id = (request.GET.get("asignatura_id") or "").strip()
    periodo_id    = (request.GET.get("periodo_id") or "").strip()

    # Validaciones básicas
    if not all(re.fullmatch(r"\d{1,10}", v or "") for v in [grupo_id, asignatura_id, periodo_id]):
        return JsonResponse({"detail": "Parámetros inválidos."}, status=400)

    with connection.cursor() as cur:
        # Notas existentes solo de estudiantes ACTIVOS en el grupo
        cur.execute("""
            SELECT n.estudiante_id, n.nota, n.fallas
            FROM public.notas n
            JOIN public.estudiante_grupo eg
              ON eg.estudiante_id = n.estudiante_id
             AND eg.grupo_id = %s
             AND eg.fecha_fin IS NULL
            WHERE n.asignatura_id = %s
              AND n.periodo_id = %s
            ORDER BY n.estudiante_id;
        """, [grupo_id, asignatura_id, periodo_id])
        notas = [{"estudiante_id": r[0], "nota": float(r[1]), "fallas": int(r[2])} for r in cur.fetchall()]

    return JsonResponse({"ok": True, "notas": notas})

# ======= NUEVO: obtener notas ya registradas para precargar en la tabla =======
@login_required(login_url="login")
@require_GET
def api_notas_por_grupo(request: HttpRequest) -> JsonResponse:
    """
    Devuelve las notas existentes para precargar la UI.
    Parámetros (GET):
      - grupo_id (obligatorio)
      - asignatura_id (obligatorio)
      - periodo_id (obligatorio)
      - estudiante_id (opcional) -> filtra por un alumno específico
    Solo retorna alumnos ACTIVOS en el grupo.
    Respuesta:
      {
        "ok": true,
        "notas": [{"estudiante_id": 1, "nota": 4.3, "fallas": 2}, ...],
        "mapa": {"1": {"nota": 4.3, "fallas": 2}, ...}
      }
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)

    grupo_id = (request.GET.get("grupo_id") or "").strip()
    asignatura_id = (request.GET.get("asignatura_id") or "").strip()
    periodo_id = (request.GET.get("periodo_id") or "").strip()
    estudiante_id = (request.GET.get("estudiante_id") or "").strip()

    if not all(re.fullmatch(r"\d{1,10}", v or "") for v in [grupo_id, asignatura_id, periodo_id]):
        return JsonResponse({"detail": "Parámetros inválidos."}, status=400)
    if estudiante_id and not re.fullmatch(r"\d{1,10}", estudiante_id):
        return JsonResponse({"detail": "Parámetro 'estudiante_id' inválido."}, status=400)

    params = [asignatura_id, periodo_id, grupo_id]
    sql = """
        SELECT n.estudiante_id, n.nota, n.fallas
          FROM public.notas n
         WHERE n.asignatura_id = %s
           AND n.periodo_id    = %s
           AND EXISTS (
                 SELECT 1
                   FROM public.estudiante_grupo eg
                  WHERE eg.estudiante_id = n.estudiante_id
                    AND eg.grupo_id      = %s
                    AND eg.fecha_fin IS NULL
               )
    """
    if estudiante_id:
        sql += " AND n.estudiante_id = %s"
        params.append(estudiante_id)

    with connection.cursor() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()

    notas = [
        {"estudiante_id": r[0], "nota": float(r[1]), "fallas": int(r[2])}
        for r in rows
    ]
    mapa = {str(r["estudiante_id"]): {"nota": r["nota"], "fallas": r["fallas"]} for r in notas}

    return JsonResponse({"ok": True, "notas": notas, "mapa": mapa})

@login_required(login_url="login")
@require_GET
def api_reporte_academico_grupo(request: HttpRequest) -> JsonResponse:
    """
    Reporte por grupo (agregado por periodo):
      - Promedio de nota por estudiante (promedio de todas sus asignaturas del grupo en ese periodo).
      - Suma de fallas en el periodo.
    Parámetros:
      ?grupo_id=<id> (obligatorio)
      ?periodo_id=<id> (opcional; si se omite se toma el último periodo abierto)
    Respuesta:
      {"ok": true, "periodo_id": <int>, "filas": [
        {"estudiante_id": 1, "apellidos":"...", "nombre":"...", "nota": 4.2, "fallas": 3}, ...
      ]}
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)

    grupo_id   = (request.GET.get("grupo_id") or "").strip()
    periodo_id = (request.GET.get("periodo_id") or "").strip()

    if not re.fullmatch(r"\d{1,10}", grupo_id or ""):
        return JsonResponse({"detail": "Parámetro 'grupo_id' inválido."}, status=400)

    # Si no viene periodo_id, tomamos el último abierto
    if not periodo_id:
        with connection.cursor() as cur:
            cur.execute("SELECT id FROM public.periodos WHERE abierto=true ORDER BY id DESC LIMIT 1;")
            row = cur.fetchone()
        if not row:
            return JsonResponse({"detail": "No hay periodos abiertos."}, status=400)
        periodo_id = str(row[0])
    elif not re.fullmatch(r"\d{1,10}", periodo_id or ""):
        return JsonResponse({"detail": "Parámetro 'periodo_id' inválido."}, status=400)

    with connection.cursor() as cur:
        # Promedio de nota del periodo y suma de fallas por estudiante ACTIVO en el grupo
        cur.execute("""
            SELECT e.id                                   AS estudiante_id,
                   e.apellidos,
                   e.nombre,
                   ROUND(AVG(n.nota)::numeric, 2)         AS nota_prom,
                   COALESCE(SUM(n.fallas), 0)             AS fallas_tot
              FROM public.estudiante_grupo eg
              JOIN public.estudiantes e
                ON e.id = eg.estudiante_id
              LEFT JOIN public.notas n
                ON n.estudiante_id = e.id
               AND n.periodo_id = %s
               AND n.asignatura_id IN (
                     SELECT asignatura_id
                       FROM public.grupo_asignatura
                      WHERE grupo_id = %s
                   )
             WHERE eg.grupo_id = %s
               AND eg.fecha_fin IS NULL
             GROUP BY e.id, e.apellidos, e.nombre
             ORDER BY e.apellidos, e.nombre;
        """, [periodo_id, grupo_id, grupo_id])
        rows = cur.fetchall()

    filas = [{
        "estudiante_id": r[0],
        "apellidos": r[1],
        "nombre": r[2],
        "nota": float(r[3]) if r[3] is not None else None,
        "fallas": int(r[4] or 0),
    } for r in rows]

    return JsonResponse({"ok": True, "periodo_id": int(periodo_id), "filas": filas})


@login_required(login_url="login")
@require_GET
def api_reporte_academico_estudiante(request: HttpRequest) -> JsonResponse:
    """
    Reporte por estudiante (agregado por periodo) dentro de un grupo:
      - Promedio de nota del periodo (todas las asignaturas del grupo).
      - Suma de fallas del periodo.
    Parámetros:
      ?grupo_id=<id> (obligatorio)
      ?estudiante_id=<id> (obligatorio y debe estar ACTIVO en ese grupo)
      ?periodo_id=<id> (opcional; si se omite se toma el último periodo abierto)
    Respuesta:
      {"ok": true, "periodo_id": <int>, "fila": {
        "estudiante_id": 1, "apellidos":"...", "nombre":"...", "nota": 4.5, "fallas": 1
      }}
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"detail": "No autorizado."}, status=403)

    grupo_id      = (request.GET.get("grupo_id") or "").strip()
    estudiante_id = (request.GET.get("estudiante_id") or "").strip()
    periodo_id    = (request.GET.get("periodo_id") or "").strip()

    if not re.fullmatch(r"\d{1,10}", grupo_id or ""):
        return JsonResponse({"detail": "Parámetro 'grupo_id' inválido."}, status=400)
    if not re.fullmatch(r"\d{1,10}", estudiante_id or ""):
        return JsonResponse({"detail": "Parámetro 'estudiante_id' inválido."}, status=400)

    # periodo por defecto: último abierto
    if not periodo_id:
        with connection.cursor() as cur:
            cur.execute("SELECT id FROM public.periodos WHERE abierto=true ORDER BY id DESC LIMIT 1;")
            row = cur.fetchone()
        if not row:
            return JsonResponse({"detail": "No hay periodos abiertos."}, status=400)
        periodo_id = str(row[0])
    elif not re.fullmatch(r"\d{1,10}", periodo_id or ""):
        return JsonResponse({"detail": "Parámetro 'periodo_id' inválido."}, status=400)

    with connection.cursor() as cur:
        # Debe estar activo en el grupo
        cur.execute("""
            SELECT 1
              FROM public.estudiante_grupo
             WHERE estudiante_id = %s
               AND grupo_id      = %s
               AND fecha_fin IS NULL;
        """, [estudiante_id, grupo_id])
        if cur.fetchone() is None:
            return JsonResponse({"detail": "El estudiante no pertenece (activo) a ese grupo."}, status=400)

        # Agregado del periodo
        cur.execute("""
            SELECT e.id, e.apellidos, e.nombre,
                   ROUND(AVG(n.nota)::numeric, 2) AS nota_prom,
                   COALESCE(SUM(n.fallas), 0)     AS fallas_tot
              FROM public.estudiantes e
              LEFT JOIN public.notas n
                ON n.estudiante_id = e.id
               AND n.periodo_id    = %s
               AND n.asignatura_id IN (
                     SELECT asignatura_id
                       FROM public.grupo_asignatura
                      WHERE grupo_id = %s
                   )
             WHERE e.id = %s
             GROUP BY e.id, e.apellidos, e.nombre
             LIMIT 1;
        """, [periodo_id, grupo_id, estudiante_id])
        r = cur.fetchone()

    fila = None
    if r:
        fila = {
            "estudiante_id": r[0],
            "apellidos": r[1],
            "nombre": r[2],
            "nota": float(r[3]) if r[3] is not None else None,
            "fallas": int(r[4] or 0),
        }

    return JsonResponse({"ok": True, "periodo_id": int(periodo_id), "fila": fila})

# ====== GUARDAR NOTAS (insert/update + historial) ======
@login_required(login_url="login")
@require_POST
@csrf_exempt
def rector_notas_guardar(request: HttpRequest) -> JsonResponse:
    """
    Recibe JSON:
    {
      "grupo_id": ..., "asignatura_id": ..., "periodo_id": ...,
      "filas": [{"estudiante_id":..., "nota":"4.3"|"", "fallas":"2"|""}, ...]
    }
    Reglas:
      - nota vacía -> 1.0; fuera de [1.0,5.0] -> error.
      - fallas vacía -> 0; negativa -> 0.
      - UPSERT en public.notas (y deja rastro en notas_historial).
    Validaciones añadidas:
      - periodo debe estar ABIERTO.
      - cada estudiante DEBE estar ACTIVO en el grupo.
      - se deduplican filas por estudiante (última gana).
    """
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"ok": False, "msg": "No autorizado."}, status=403)

    # Parse
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "msg": "JSON inválido."}, status=400)

    grupo_id      = str(payload.get("grupo_id") or "").strip()
    asignatura_id = str(payload.get("asignatura_id") or "").strip()
    periodo_id    = str(payload.get("periodo_id") or "").strip()
    filas         = payload.get("filas") or []

    # Básicas
    if not all(re.fullmatch(r"\d{1,10}", v or "") for v in [grupo_id, asignatura_id, periodo_id]):
        return JsonResponse({"ok": False, "msg": "Parámetros inválidos."}, status=400)

    # 1) Periodo abierto
    if not _periodo_abierto(periodo_id):
        return JsonResponse({"ok": False, "msg": "El periodo no está abierto para registrar notas."}, status=400)

    # 2) Validar que la asignatura pertenezca al grupo
    with connection.cursor() as cur:
        cur.execute("""
            SELECT 1
            FROM public.grupo_asignatura
            WHERE grupo_id=%s AND asignatura_id=%s;
        """, [grupo_id, asignatura_id])
        if cur.fetchone() is None:
            return JsonResponse({"ok": False, "msg": "La asignatura no está ligada a ese grupo."}, status=400)

    # 3) Cargar estudiantes activos del grupo y normalizar filas
    activos = _estudiantes_activos_en_grupo(grupo_id)
    if not activos:
        return JsonResponse({"ok": False, "msg": "El grupo no tiene estudiantes activos."}, status=400)

    # Deduplicar por estudiante (última fila para cada uno)
    normalizadas: dict[str, tuple[Decimal, int]] = {}
    fuera_del_grupo: list[str] = []

    for f in filas:
        est_id = str(f.get("estudiante_id") or "").strip()
        if not re.fullmatch(r"\d{1,10}", est_id or ""):
            return JsonResponse({"ok": False, "msg": "ID de estudiante inválido."}, status=400)

        if est_id not in activos:
            fuera_del_grupo.append(est_id)
            continue

        raw_nota   = (f.get("nota") or "").strip()
        raw_fallas = (f.get("fallas") or "").strip()

        try:
            nota = Decimal(raw_nota) if raw_nota != "" else Decimal("1.0")
        except InvalidOperation:
            return JsonResponse({"ok": False, "msg": "Nota inválida."}, status=400)
        if nota < Decimal("1.0") or nota > Decimal("5.0"):
            return JsonResponse({"ok": False, "msg": "Las notas deben estar entre 1.0 y 5.0."}, status=400)

        try:
            fallas = int(raw_fallas) if raw_fallas != "" else 0
        except ValueError:
            return JsonResponse({"ok": False, "msg": "Fallas inválidas."}, status=400)
        fallas = max(0, fallas)

        normalizadas[est_id] = (nota, fallas)  # última gana

    if fuera_del_grupo:
        # Si prefieres ignorarlos silenciosamente, elimina este return y sigue;
        # por defecto devolvemos error explícito.
        return JsonResponse(
            {"ok": False,
             "msg": f"Estudiantes no activos en el grupo: {', '.join(sorted(set(fuera_del_grupo)))}"},
            status=400
        )

    created = updated = 0

    with transaction.atomic():
        with connection.cursor() as cur:
            for est_id, (nota, fallas) in normalizadas.items():
                # ¿Existe?
                cur.execute("""
                    SELECT id, nota, fallas
                    FROM public.notas
                    WHERE estudiante_id=%s AND asignatura_id=%s AND periodo_id=%s;
                """, [est_id, asignatura_id, periodo_id])
                row = cur.fetchone()

                if row:
                    nota_id, nota_prev, fall_prev = row

                    # historial
                    cur.execute("""
                        INSERT INTO public.notas_historial
                          (nota_id, estudiante_id, asignatura_id, periodo_id,
                           nota_anterior, fallas_anterior, nota_nueva, fallas_nuevas,
                           accion, realizado_por_usuario)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'UPDATE',%s);
                    """, [nota_id, est_id, asignatura_id, periodo_id,
                          nota_prev, fall_prev, nota, fallas, request.user.id])

                    # update
                    cur.execute("""
                        UPDATE public.notas
                           SET nota=%s, fallas=%s, actualizado_por_usuario=%s, fuente_rol=%s, actualizado_en=now()
                         WHERE id=%s;
                    """, [nota, fallas, request.user.id, 'RECTOR', nota_id])
                    updated += 1
                else:
                    # insert
                    cur.execute("""
                        INSERT INTO public.notas
                          (estudiante_id, asignatura_id, periodo_id, nota, fallas,
                           actualizado_por_usuario, fuente_rol)
                        VALUES (%s,%s,%s,%s,%s,%s,'RECTOR')
                        RETURNING id;
                    """, [est_id, asignatura_id, periodo_id, nota, fallas, request.user.id])
                    nota_id = cur.fetchone()[0]

                    # historial
                    cur.execute("""
                        INSERT INTO public.notas_historial
                          (nota_id, estudiante_id, asignatura_id, periodo_id,
                           nota_nueva, fallas_nuevas, accion, realizado_por_usuario)
                        VALUES (%s,%s,%s,%s,%s,%s,'INSERT',%s);
                    """, [nota_id, est_id, asignatura_id, periodo_id, nota, fallas, request.user.id])
                    created += 1

    return JsonResponse({"ok": True, "msg": "Notas guardadas.", "creadas": created, "actualizadas": updated})

# ====== EXPORTES ======
@login_required(login_url="login")
@require_GET
def export_notas_excel(request: HttpRequest) -> HttpResponse:
    """Genera Excel para los parámetros dados."""
    if (resp := _guard_rector(request)) is not None:
        return redirect("login")

    grupo_id = request.GET.get("grupo_id")
    asignatura_id = request.GET.get("asignatura_id")
    periodo_id = request.GET.get("periodo_id")
    if not all(re.fullmatch(r"\d{1,10}", v or "") for v in [grupo_id, asignatura_id, periodo_id]):
        return HttpResponse("Parámetros inválidos.", status=400)

    # Encabezado (sede, grado, grupo, área, asignatura, periodo, docente)
    with connection.cursor() as cur:
        cur.execute("""
            SELECT s.nombre, gr.nombre, g.nombre, ar.nombre, a.nombre, p.nombre,
                   COALESCE(u.nombre || ' ' || u.apellidos, 'N/A') AS docente
            FROM public.grupos g
            JOIN public.sedes s ON s.id = g.sede_id
            JOIN public.grados gr ON gr.id = g.grado_id
            JOIN public.grupo_asignatura ga ON ga.grupo_id = g.id
            JOIN public.asignaturas a ON a.id = ga.asignatura_id
            JOIN public.areas ar ON ar.id = a.area_id
            JOIN public.periodos p ON p.id = %s
            LEFT JOIN public.docente_asignacion da ON da.grupo_asignatura_id = ga.id AND a.id=%s
            LEFT JOIN public.docentes d ON d.id = da.docente_id
            LEFT JOIN public.usuarios u ON u.id = d.usuario_id
            WHERE g.id=%s AND a.id=%s
            LIMIT 1;
        """, [periodo_id, asignatura_id, grupo_id, asignatura_id])
        meta = cur.fetchone()

        cur.execute("""
            SELECT e.apellidos, e.nombre,
                   COALESCE(n.nota, 1.0) AS nota,
                   COALESCE(n.fallas, 0) AS fallas
            FROM public.estudiante_grupo eg
            JOIN public.estudiantes e ON e.id = eg.estudiante_id
            LEFT JOIN public.notas n ON n.estudiante_id = e.id
                 AND n.asignatura_id = %s AND n.periodo_id = %s
            WHERE eg.grupo_id = %s AND eg.fecha_fin IS NULL
            ORDER BY e.apellidos, e.nombre;
        """, [asignatura_id, periodo_id, grupo_id])
        filas = cur.fetchall()

    # Excel con openpyxl
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse("Falta dependencia 'openpyxl'.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas"

    ws.append(["Sede","Grado","Grupo","Área","Asignatura","Periodo","Docente"])
    if meta:
        ws.append(list(meta))
    ws.append([])
    ws.append(["APELLIDOS","NOMBRES","NOTA","FALLAS"])
    for r in filas:
        ws.append(list(r))

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf); buf.seek(0)

    resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp['Content-Disposition'] = 'attachment; filename="notas_grupo.xlsx"'
    return resp

@login_required(login_url="login")
@require_GET
def export_notas_pdf(request: HttpRequest) -> HttpResponse:
    """PDF sencillo usando reportlab."""
    if (resp := _guard_rector(request)) is not None:
        return redirect("login")

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("Falta dependencia 'reportlab'.", status=500)

    grupo_id = request.GET.get("grupo_id")
    asignatura_id = request.GET.get("asignatura_id")
    periodo_id = request.GET.get("periodo_id")
    if not all(re.fullmatch(r"\d{1,10}", v or "") for v in [grupo_id, asignatura_id, periodo_id]):
        return HttpResponse("Parámetros inválidos.", status=400)

    with connection.cursor() as cur:
        cur.execute("""
            SELECT s.nombre, gr.nombre, g.nombre, ar.nombre, a.nombre, p.nombre,
                   COALESCE(u.nombre || ' ' || u.apellidos, 'N/A') AS docente
            FROM public.grupos g
            JOIN public.sedes s ON s.id = g.sede_id
            JOIN public.grados gr ON gr.id = g.grado_id
            JOIN public.grupo_asignatura ga ON ga.grupo_id = g.id
            JOIN public.asignaturas a ON a.id = ga.asignatura_id
            JOIN public.areas ar ON ar.id = a.area_id
            JOIN public.periodos p ON p.id = %s
            LEFT JOIN public.docente_asignacion da ON da.grupo_asignatura_id = ga.id AND a.id=%s
            LEFT JOIN public.docentes d ON d.id = da.docente_id
            LEFT JOIN public.usuarios u ON u.id = d.usuario_id
            WHERE g.id=%s AND a.id=%s
            LIMIT 1;
        """, [periodo_id, asignatura_id, grupo_id, asignatura_id])
        meta = cur.fetchone()

        cur.execute("""
            SELECT e.apellidos, e.nombre,
                   COALESCE(n.nota, 1.0) AS nota,
                   COALESCE(n.fallas, 0) AS fallas
            FROM public.estudiante_grupo eg
            JOIN public.estudiantes e ON e.id = eg.estudiante_id
            LEFT JOIN public.notas n ON n.estudiante_id = e.id
                 AND n.asignatura_id = %s AND n.periodo_id = %s
            WHERE eg.grupo_id = %s AND eg.fecha_fin IS NULL
            ORDER BY e.apellidos, e.nombre;
        """, [asignatura_id, periodo_id, grupo_id])
        filas = cur.fetchall()

    from io import BytesIO
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    y = h - 2*cm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y, "Registro de Notas — Por grupo"); y -= 0.8*cm
    c.setFont("Helvetica", 10)
    if meta:
        s, gr, g, ar, a, p, d = meta
        for line in [f"Sede: {s}", f"Grado: {gr}", f"Grupo: {g}", f"Área: {ar}",
                     f"Asignatura: {a}", f"Periodo: {p}", f"Docente: {d}"]:
            c.drawString(2*cm, y, line); y -= 0.55*cm
    y -= 0.3*cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y, "APELLIDOS"); c.drawString(8*cm, y, "NOMBRES")
    c.drawString(13.5*cm, y, "NOTA"); c.drawString(16*cm, y, "FALLAS")
    y -= 0.4*cm
    c.setFont("Helvetica", 10)
    for ap, no, nt, fa in filas:
      if y < 2*cm: c.showPage(); y = h - 2*cm
      c.drawString(2*cm, y, str(ap)[:35]); c.drawString(8*cm, y, str(no)[:28])
      c.drawString(13.5*cm, y, str(nt)); c.drawString(16*cm, y, str(fa))
      y -= 0.38*cm

    c.showPage(); c.save(); buf.seek(0)
    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="notas_grupo.pdf"'
    return resp

# ===== Helpers de boletines (no tocan lógica previa) =====
def _nivel_desempeno_rango(nota: float | None) -> str:
    if nota is None:
        return ""
    # Rangos del instructivo:
    # 4.5–5.0 Superior, 4.0–4.4 Alto, 3.5–3.9 Básico, 1.0–3.4 Bajo
    if 4.5 <= nota <= 5.0:
        return "SUPERIOR"
    if 4.0 <= nota <= 4.4:
        return "ALTO"
    if 3.5 <= nota <= 3.9:
        return "BÁSICO"
    if 1.0 <= nota <= 3.4:
        return "BAJO"
    return ""

def _nivel_desempeno(nota: float | None) -> str:
    # Alias para mantener compatibilidad con el código existente
    return _nivel_desempeno_rango(nota)

    
def _ponderaciones_por_grado(grado_nombre: str) -> dict[str, float]:
    """
    (Opcional) Ponderaciones de P1/P2/P3 para un grado.
    Si no hay regla para ese grado, usa 1/3 cada uno.
    """
    reglas = {
        # Ejemplos (ajústalos si quieres):
        # "Primaria": {"p1": 0.20, "p2": 0.30, "p3": 0.50},
        # "Bachillerato": {"p1": 0.33, "p2": 0.33, "p3": 0.34},
        # "10°": {"p1": 0.25, "p2": 0.35, "p3": 0.40},
    }
    pesos = reglas.get(grado_nombre, {"p1": 1/3, "p2": 1/3, "p3": 1/3})
    total = (pesos.get("p1", 0) + pesos.get("p2", 0) + pesos.get("p3", 0)) or 1.0
    return {
        "p1": float(pesos["p1"] / total),
        "p2": float(pesos["p2"] / total),
        "p3": float(pesos["p3"] / total),
    }

def _prom_ponderado(p1: float | None, p2: float | None, p3: float | None,
                    w: dict[str, float]) -> float | None:
    """
    Promedio ponderado usando sólo los periodos presentes.
    Re-normaliza los pesos a los presentes.
    """
    pares = []
    if p1 is not None: pares.append(("p1", p1))
    if p2 is not None: pares.append(("p2", p2))
    if p3 is not None: pares.append(("p3", p3))
    if not pares:
        return None
    suma_pesos = sum(w[k] for k, _ in pares) or 1.0
    return round(sum(val * (w[k] / suma_pesos) for k, val in pares), 2)


def _es_perdida(nombre_asig_o_area: str, nota: float | None) -> bool:
    """
    Determina si una nota es pérdida.
    Regla general: pierde < 3.0
    Regla especial: en Inglés pierde < 3.5
    """
    if nota is None:
        return False
    umbral = 3.0
    # Regla especial para Inglés (cubre nombres como 'INGLES', 'HUMANIDADES INGLES', etc.)
    if "INGLES" in (nombre_asig_o_area or "").upper():
        umbral = 3.5
    return nota < umbral


def _dense_rank_desc(valores: list[tuple[int, float]]) -> dict[int, int]:
    """
    Recibe lista [(est_id, promedio), ...] y retorna {est_id: puesto}
    Empates comparten el mismo puesto (dense rank).
    """
    # Ordenar desc por promedio
    orden = sorted(valores, key=lambda x: (-x[1], x[0]))
    puestos = {}
    puesto_actual = 0
    previo = None
    for idx, (est, prom) in enumerate(orden, 1):
        if previo is None or prom != previo:
            puesto_actual += 1
            previo = prom
        puestos[est] = puesto_actual
    return puestos

def _cargar_meta_grupo(grupo_id: str, periodo_id: str):
    with connection.cursor() as cur:
        cur.execute("""
            SELECT s.nombre AS sede, gr.nombre AS grado, g.nombre AS grupo
            FROM public.grupos g
            JOIN public.sedes s  ON s.id  = g.sede_id
            JOIN public.grados gr ON gr.id = g.grado_id
            WHERE g.id = %s
        """, [grupo_id])
        sede, grado, grupo = cur.fetchone()

        cur.execute("SELECT nombre FROM public.periodos WHERE id=%s", [periodo_id])
        per = cur.fetchone()
        periodo_nombre = per[0] if per else f"ID {periodo_id}"
    return {
        "sede": sede, "grado": grado, "grupo": grupo,
        "periodo_nombre": periodo_nombre, "anio": datetime.now().year,
        "jornada": "ÚNICA", "colegio": "Nombre del Colegio",
        "fecha_emision": datetime.now().strftime("%Y-%m-%d"),
    }

def _periodo_abierto(periodo_id: str) -> bool:
    """True si el periodo está abierto."""
    with connection.cursor() as cur:
        cur.execute("SELECT abierto FROM public.periodos WHERE id=%s;", [periodo_id])
        row = cur.fetchone()
    return bool(row and row[0])

def _estudiantes_activos_en_grupo(grupo_id: str) -> set[str]:
    """Conjunto de IDs (str) de estudiantes ACTIVOS en el grupo."""
    with connection.cursor() as cur:
        cur.execute("""
            SELECT e.id
            FROM public.estudiante_grupo eg
            JOIN public.estudiantes e ON e.id = eg.estudiante_id
            WHERE eg.grupo_id=%s AND eg.fecha_fin IS NULL
        """, [grupo_id])
        return {str(r[0]) for r in cur.fetchall()}

def _dataset_boletines(grupo_id: str, periodo_id: str, estudiante_id: str | None = None):
    """
    Construye el dataset para renderizar boletines:
      - Lista de estudiantes (activos) del grupo (filtrando si 'estudiante_id' llega).
      - Áreas y asignaturas del grupo.
      - Notas por P1/P2/P3 y Final; fallas por periodo seleccionado.
      - Promedio y puestos (grupo, grado, institucional).
    """
    # 1) Estudiantes activos del grupo
    with connection.cursor() as cur:
        params = [grupo_id]
        sql_est = """
            SELECT e.id, e.apellidos, e.nombre, e.documento
            FROM public.estudiante_grupo eg
            JOIN public.estudiantes e ON e.id = eg.estudiante_id
            WHERE eg.grupo_id=%s AND eg.fecha_fin IS NULL
        """
        if estudiante_id:
            sql_est += " AND e.id=%s"
            params.append(estudiante_id)
        sql_est += " ORDER BY e.apellidos, e.nombre"
        cur.execute(sql_est, params)
        estudiantes = [{"id": r[0], "apellidos": r[1], "nombres": r[2], "documento": r[3]} for r in cur.fetchall()]
        if not estudiantes:
            return [], [], {}

        est_ids = [e["id"] for e in estudiantes]

        # 2) Áreas y asignaturas del grupo
        cur.execute("""
            SELECT ar.id AS area_id, ar.nombre AS area, a.id AS asig_id, a.nombre AS asig
            FROM public.grupo_asignatura ga
            JOIN public.asignaturas a ON a.id = ga.asignatura_id
            JOIN public.areas ar ON ar.id = a.area_id
            WHERE ga.grupo_id = %s
            ORDER BY ar.nombre, a.nombre
        """, [grupo_id])
        rows = cur.fetchall()
        areas = []  # [(area_id, area_nombre, [ (asig_id, asig_nombre), ... ])]
        tmp = defaultdict(list)
        area_nombres = {}
        for area_id, area, asig_id, asig in rows:
            tmp[area_id].append((asig_id, asig))
            area_nombres[area_id] = area
        for aid, asigns in tmp.items():
            areas.append((aid, area_nombres[aid], asigns))

        # 3) Notas de P1..P3 para todas las asignaturas de estos estudiantes
        cur.execute("""
            SELECT n.estudiante_id, n.asignatura_id, n.periodo_id, n.nota, n.fallas
            FROM public.notas n
            WHERE n.estudiante_id = ANY(%s)
              AND n.asignatura_id IN (
                    SELECT a.id FROM public.grupo_asignatura ga
                    JOIN public.asignaturas a ON a.id = ga.asignatura_id
                    WHERE ga.grupo_id = %s
              )
        """, [est_ids, grupo_id])
        notas_rows = cur.fetchall()

        # (Ponderaciones por grado) — obtenemos el nombre del grado del grupo
        cur.execute("""
            SELECT gr.nombre
            FROM public.grupos g
            JOIN public.grados gr ON gr.id = g.grado_id
            WHERE g.id=%s
        """, [grupo_id])
        grado_nombre = (cur.fetchone() or [""])[0]

    # === Punto C: pesos por grado (opcional) ===
    # Si existe el helper global _ponderaciones_por_grado, úsalo; si no, 1/3 cada uno.
    try:
        pesos = _ponderaciones_por_grado(grado_nombre)  # type: ignore[name-defined]
    except NameError:
        pesos = {"p1": 1/3, "p2": 1/3, "p3": 1/3}

    # Promedio ponderado usando sólo los periodos presentes (re-normaliza pesos)
    def _prom_ponderado(p1: float | None, p2: float | None, p3: float | None) -> float | None:
        pares = []
        if p1 is not None: pares.append(("p1", p1))
        if p2 is not None: pares.append(("p2", p2))
        if p3 is not None: pares.append(("p3", p3))
        if not pares:
            return None
        suma_pesos = sum(pesos[k] for k, _ in pares) or 1.0
        return round(sum(val * (pesos[k] / suma_pesos) for k, val in pares), 2)

    # Estructurar: notas[(est, asig)]['p1'|'p2'|'p3'|'final'|'fallas_p?']
    notas = defaultdict(lambda: {"p1": None, "p2": None, "p3": None, "final": None,
                                 "f1": 0, "f2": 0, "f3": 0})
    for est, asig, per, nt, fa in notas_rows:
        key = (est, asig)
        if per == 1:
            notas[key]["p1"] = float(nt) if nt is not None else None
            notas[key]["f1"] = int(fa or 0)
        elif per == 2:
            notas[key]["p2"] = float(nt) if nt is not None else None
            notas[key]["f2"] = int(fa or 0)
        elif per == 3:
            notas[key]["p3"] = float(nt) if nt is not None else None
            notas[key]["f3"] = int(fa or 0)
        # final se calculará abajo con los periodos existentes (ahora ponderado)

    # 4) Construir detalle por estudiante y calcular promedio
    estudiantes_det = []
    promedios_grupo = []
    promedios_por_est = {}  # est_id -> promedio

    for e in estudiantes:
        est_id = e["id"]
        detalle = []
        area_promedios = []

        for area_id, area_nombre, asigns in areas:
            # Filas de asignaturas
            filas_asig = []
            area_p1 = []; area_p2 = []; area_p3 = []
            for asig_id, asig_nombre in asigns:
                n = notas[(est_id, asig_id)]
                # === (C) Final de la asignatura: promedio ponderado P1/P2/P3 presentes ===
                final_asig = _prom_ponderado(n["p1"], n["p2"], n["p3"])

                # fallas del periodo seleccionado
                fsel = 0
                if str(periodo_id) == "1": fsel = n["f1"]
                elif str(periodo_id) == "2": fsel = n["f2"]
                elif str(periodo_id) == "3": fsel = n["f3"]

                filas_asig.append({
                    "tipo": "ASIG",
                    "nombre": asig_nombre,
                    "p1": n["p1"] if n["p1"] is not None else "",
                    "p2": n["p2"] if n["p2"] is not None else "",
                    "p3": n["p3"] if n["p3"] is not None else "",
                    "final": final_asig if final_asig is not None else "",
                    "fallas": fsel if fsel else "",
                    "desempeno": _nivel_desempeno(
                        final_asig if str(periodo_id) == "3" else
                        (n["p2"] if str(periodo_id) == "2" else n["p1"])
                    )
                })
                if n["p1"] is not None: area_p1.append(n["p1"])
                if n["p2"] is not None: area_p2.append(n["p2"])
                if n["p3"] is not None: area_p3.append(n["p3"])

            # Fila de área (promedio de las asignaturas del área)
            def _avg(lst): return round(sum(lst)/len(lst), 2) if lst else None
            ap1 = _avg(area_p1); ap2 = _avg(area_p2); ap3 = _avg(area_p3)

            # === (C) Final del área: ponderado P1/P2/P3 presentes ===
            afinal = _prom_ponderado(ap1, ap2, ap3)

            detalle.append({
                "tipo": "AREA", "nombre": area_nombre,
                "p1": ap1 if ap1 is not None else "",
                "p2": ap2 if ap2 is not None else "",
                "p3": ap3 if ap3 is not None else "",
                "final": afinal if afinal is not None else "",
                "desempeno": _nivel_desempeno(
                    afinal if str(periodo_id) == "3" else
                    (ap2 if str(periodo_id) == "2" else ap1)
                ),
            })
            detalle.extend(filas_asig)
            if afinal is not None:
                area_promedios.append(afinal)

        # Promedio general del estudiante (promedio de áreas con valor)
        prom = round(sum(area_promedios)/len(area_promedios), 2) if area_promedios else 0.0
        promedios_grupo.append((est_id, prom))
        promedios_por_est[est_id] = prom

        estudiantes_det.append({
            **e,
            "detalle": detalle,
            "promedio": prom,
        })

    # 5) Ranking (grupo)
    puestos_grupo = _dense_rank_desc(promedios_grupo)

    # 6) Ranking grado e institucional
    with connection.cursor() as cur:
        # grado_id del grupo
        cur.execute("SELECT grado_id FROM public.grupos WHERE id=%s", [grupo_id])
        grado_id = cur.fetchone()[0]

        # Todos los estudiantes activos del grado (todas las divisiones)
        cur.execute("""
            SELECT e.id
            FROM public.estudiante_grupo eg
            JOIN public.grupos g ON g.id = eg.grupo_id
            JOIN public.estudiantes e ON e.id = eg.estudiante_id
            WHERE eg.fecha_fin IS NULL AND g.grado_id=%s
        """, [grado_id])
        est_grado_ids = [r[0] for r in cur.fetchall()]
        if not est_grado_ids:
            est_grado_ids = []

    def _promedio_estudiante_fuera(est_id: int) -> float:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT ar.id, AVG(finals.nf) AS area_final
                FROM public.areas ar
                JOIN public.asignaturas a ON a.area_id = ar.id
                JOIN (
                    SELECT n.asignatura_id, AVG(n.nota) AS nf
                    FROM public.notas n
                    WHERE n.estudiante_id=%s
                    GROUP BY n.asignatura_id
                ) finals ON finals.asignatura_id = a.id
                GROUP BY ar.id
            """, [est_id])
            arr = [float(r[1]) for r in cur.fetchall()]
            return round(sum(arr)/len(arr), 2) if arr else 0.0

    proms_grado = []
    for eid in est_grado_ids:
        proms_grado.append((eid, promedios_por_est.get(eid) or _promedio_estudiante_fuera(eid)))
    puestos_grado = _dense_rank_desc(proms_grado) if proms_grado else {}

    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT e.id
            FROM public.estudiante_grupo eg
            JOIN public.estudiantes e ON e.id = eg.estudiante_id
            WHERE eg.fecha_fin IS NULL
        """)
        est_inst_ids = [r[0] for r in cur.fetchall()]
    proms_inst = []
    for eid in est_inst_ids:
        proms_inst.append((eid, promedios_por_est.get(eid) or _promedio_estudiante_fuera(eid)))
    puestos_inst = _dense_rank_desc(proms_inst) if proms_inst else {}

    for s in estudiantes_det:
        eid = s["id"]
        s["puesto_grupo"] = puestos_grupo.get(eid, "")
        s["puesto_grado"] = puestos_grado.get(eid, "")
        s["puesto_institucional"] = puestos_inst.get(eid, "")

    meta = {}  # la vista principal añadirá _cargar_meta_grupo()
    return estudiantes_det, areas, meta

@login_required(login_url="login")
@require_GET
def rector_reportes_academicos_export(request):
    """
    Exporta boletines:
      - del GRUPO si llega ?grupo_id=
      - o del GRADO (en una sede) si llegan ?sede_id=&grado_id= y NO llega grupo_id
      - opcionalmente de un solo estudiante si llega ?estudiante_id=
    Requiere ?periodo_id= y ?formato=pdf|excel
    """
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.conf import settings
    # Import compat de WeasyPrint (no rompe en local sin DLLs)
    from core.utils.weasy_compat import HTML, CSS, WEASY_AVAILABLE

    formato       = (request.GET.get('formato') or 'pdf').lower()
    grupo_id      = (request.GET.get('grupo_id') or '').strip()
    grado_id      = (request.GET.get('grado_id') or '').strip()
    sede_id       = (request.GET.get('sede_id') or '').strip()
    periodo_id    = (request.GET.get('periodo_id') or '').strip()
    estudiante_id = (request.GET.get('estudiante_id') or '').strip() or None  # opcional

    if not periodo_id:
        return HttpResponse("Falta periodo_id.", status=400)

    # Usa tu builder tal cual (NO se toca la lógica)
    boletines = build_boletines(
        grupo_id=grupo_id,
        grado_id=grado_id,
        sede_id=sede_id,
        periodo_id=periodo_id,
        estudiante_id=estudiante_id,
    )
    if not boletines:
        return HttpResponse("No hay información para generar boletines con los parámetros dados.", status=404)

    colegio = {
        "nombre": "IE Departamental Gustavo Uribe Ramírez",
        "subtitulo": "Secretaría de Educación de Cundinamarca – Granada",
    }
    fecha_generacion = now().strftime("%d/%m/%Y")

    # ========= EXCEL (HTML con Content-Type de Excel) =========
    if formato in ('excel', 'xlsx', 'xls'):
        html = render_to_string(
            "boletines/boletin_alumno.html",
            {
                "boletines": boletines,  # la plantilla ya itera
                "colegio": colegio,
                "fecha_generacion": fecha_generacion,
            },
            request=request,
        )
        resp = HttpResponse(html, content_type="application/vnd.ms-excel; charset=utf-8")

        if estudiante_id and len(boletines) == 1:
            filename = f"boletin_estudiante_{estudiante_id}.xls"
        elif grupo_id:
            filename = f"boletines_grupo_{grupo_id}.xls"
        elif grado_id and sede_id:
            filename = f"boletines_grado_{grado_id}_sede_{sede_id}.xls"
        else:
            filename = "boletines.xls"

        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    # ========= PDF (WeasyPrint) =========
    # Fallback automático en local si no hay DLLs de WeasyPrint → entrega Excel
    if formato == "pdf" and not WEASY_AVAILABLE:
        qs = request.GET.copy()
        qs["formato"] = "excel"
        # Redirige a la misma URL pero pidiendo Excel (evita 500 y mantiene UX)
        return redirect(f"{request.path}?{qs.urlencode()}")

    html = render_to_string(
        "boletines/boletin_alumno.html",
        {
            "boletines": boletines,
            "colegio": colegio,
            "fecha_generacion": fecha_generacion,
        },
        request=request,
    )

    pdf = HTML(
        string=html,
        base_url=request.build_absolute_uri("/")  # para resolver {% static %} e imágenes
    ).write_pdf(
        stylesheets=[
            CSS(filename=str((settings.BASE_DIR / "static" / "boletin" / "boletin.css").resolve()))
        ]
    )

    if estudiante_id and len(boletines) == 1:
        filename = f"boletin_estudiante_{estudiante_id}.pdf"
    elif grupo_id:
        filename = f"boletines_grupo_{grupo_id}.pdf"
    elif grado_id and sede_id:
        filename = f"boletines_grado_{grado_id}_sede_{sede_id}.pdf"
    else:
        filename = "boletines.pdf"

    response = HttpResponse(pdf, content_type="application/pdf")
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response



# ===== POST: Registrar / actualizar NOTA =====
@login_required(login_url="login")
@require_POST
def rector_registrar_nota(request: HttpRequest) -> JsonResponse:
    # Solo rector:
    if (resp := _guard_rector(request)) is not None:
        return JsonResponse({"ok": False, "msg": "No autorizado."}, status=403)

    # --------- INPUTS ---------
    estudiante_id = (request.POST.get("estudiante_id") or "").strip()
    asignatura_id = (request.POST.get("asignatura_id") or "").strip()
    periodo_id    = (request.POST.get("periodo_id") or "").strip()
    nota_str      = (request.POST.get("nota") or "").strip()
    fallas_str    = (request.POST.get("fallas") or "0").strip()

    # --------- VALIDACIONES BÁSICAS ---------
    # ids numéricos razonables
    if not all(re.fullmatch(r"\d{1,10}", x or "") for x in (estudiante_id, asignatura_id, periodo_id)):
        return JsonResponse({"ok": False, "msg": "Parámetros inválidos."}, status=400)

    # fallas entero >= 0
    try:
        fallas = int(fallas_str)
        if fallas < 0:
            raise ValueError()
    except ValueError:
        return JsonResponse({"ok": False, "msg": "Las fallas deben ser un número entero ≥ 0."}, status=400)

    # nota Decimal con 2 decimales entre 1.00 y 5.00 (el CHECK BD también lo refuerza)
    try:
        nota = Decimal(nota_str).quantize(Decimal("0.01"))
        if nota < Decimal("1.00") or nota > Decimal("5.00"):
            raise InvalidOperation()
    except (InvalidOperation, ValueError):
        return JsonResponse({"ok": False, "msg": "La nota debe estar entre 1.00 y 5.00 con dos decimales."}, status=400)

    # --------- LÓGICA EN BD ---------
    user_id = request.user.id  # para auditoría

    # Normalizamos el rol a MAYÚSCULAS, por compatibilidad con el enum rol_usuario
    role = getattr(request.user, "rol", None)
    fuente_rol = (str(role) if role is not None else "RECTOR").upper()

    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                # 1) Periodo debe estar ABIERTO
                cur.execute(
                    "SELECT 1 FROM public.periodos WHERE id=%s AND abierto IS TRUE;",
                    [periodo_id],
                )
                if cur.fetchone() is None:
                    return JsonResponse({"ok": False, "msg": "No se puede registrar nota: el periodo no está abierto."}, status=400)

                # 2) Verificar grupo ACTIVO del estudiante
                cur.execute("""
                    SELECT eg.grupo_id
                    FROM public.estudiante_grupo eg
                    WHERE eg.estudiante_id=%s AND eg.fecha_fin IS NULL
                    LIMIT 1;
                """, [estudiante_id])
                row_grp = cur.fetchone()
                if row_grp is None:
                    return JsonResponse({"ok": False, "msg": "El estudiante no tiene un grupo activo."}, status=400)
                grupo_id = row_grp[0]

                # 3) Asignatura ∈ grupo activo
                cur.execute("""
                    SELECT 1
                    FROM public.grupo_asignatura ga
                    WHERE ga.grupo_id=%s AND ga.asignatura_id=%s
                    LIMIT 1;
                """, [grupo_id, asignatura_id])
                if cur.fetchone() is None:
                    return JsonResponse({"ok": False, "msg": "La asignatura no pertenece al grupo activo del estudiante."}, status=400)

                # 4) ¿Existe ya la nota? (para historial)
                cur.execute("""
                    SELECT id, nota, fallas
                    FROM public.notas
                    WHERE estudiante_id=%s AND asignatura_id=%s AND periodo_id=%s
                    FOR UPDATE;
                """, [estudiante_id, asignatura_id, periodo_id])
                prev = cur.fetchone()

                if prev:
                    nota_id_ant, nota_ant, fallas_ant = prev
                    # UPDATE
                    cur.execute("""
                        UPDATE public.notas
                           SET nota=%s,
                               fallas=%s,
                               actualizado_por_usuario=%s,
                               fuente_rol=%s,
                               actualizado_en=now()
                         WHERE id=%s
                         RETURNING id;
                    """, [str(nota), fallas, user_id, fuente_rol, nota_id_ant])
                    nota_id = cur.fetchone()[0]
                    accion = 'UPDATE'
                    # Historial
                    cur.execute("""
                        INSERT INTO public.notas_historial
                            (nota_id, estudiante_id, asignatura_id, periodo_id,
                             nota_anterior, fallas_anterior,
                             nota_nueva, fallas_nuevas,
                             accion, realizado_por_usuario)
                        VALUES (%s,%s,%s,%s, %s,%s, %s,%s, %s, %s);
                    """, [nota_id, estudiante_id, asignatura_id, periodo_id,
                          nota_ant, fallas_ant, str(nota), fallas, accion, user_id])
                else:
                    # INSERT
                    cur.execute("""
                        INSERT INTO public.notas
                            (estudiante_id, asignatura_id, periodo_id, nota, fallas,
                             actualizado_por_usuario, fuente_rol, actualizado_en)
                        VALUES (%s,%s,%s, %s,%s, %s,%s, now())
                        RETURNING id;
                    """, [estudiante_id, asignatura_id, periodo_id, str(nota), fallas,
                          user_id, fuente_rol])
                    nota_id = cur.fetchone()[0]
                    accion = 'INSERT'
                    # Historial
                    cur.execute("""
                        INSERT INTO public.notas_historial
                            (nota_id, estudiante_id, asignatura_id, periodo_id,
                             nota_anterior, fallas_anterior,
                             nota_nueva, fallas_nuevas,
                             accion, realizado_por_usuario)
                        VALUES (%s,%s,%s,%s, NULL,NULL, %s,%s, %s, %s);
                    """, [nota_id, estudiante_id, asignatura_id, periodo_id,
                          str(nota), fallas, accion, user_id])

        return JsonResponse({
            "ok": True,
            "msg": "Nota registrada." if accion == "INSERT" else "Nota actualizada.",
            "accion": accion,
            "nota_id": nota_id,
        })

    except Exception as e:
        # Acepta que el helper pueda devolver más de 4 valores
        ok, msg, code, http, *_ = map_db_error(e)
        if code in ("23514",):  # check_violation (rango nota, fallas>=0, etc.)
            msg = msg or "Datos fuera de rango."
        return JsonResponse({"ok": False, "msg": msg or "No se pudo registrar la nota."}, status=http or 400)

def _estudiantes_para_exportar(*, grupo_id: str | None, sede_id: str | None, grado_id: str | None) -> list[dict]:
    """
    Devuelve los estudiantes a exportar usando SQL crudo:
      - Si llega grupo_id: todos los estudiantes ACTIVOS de ese grupo.
      - Si NO llega grupo_id pero sí sede_id y grado_id: todos los ACTIVOS de ese grado en la sede.
    Retorna: [{"id":..., "apellidos":"...", "nombre":"...", "documento":"..."}, ...]
    """
    res: list[dict] = []
    grupo_id  = (grupo_id or "").strip()
    sede_id   = (sede_id or "").strip()
    grado_id  = (grado_id or "").strip()

    with connection.cursor() as cur:
        if grupo_id:
            cur.execute(
                """
                SELECT DISTINCT e.id, e.apellidos, e.nombre, e.documento
                FROM public.estudiante_grupo eg
                JOIN public.estudiantes e ON e.id = eg.estudiante_id
                WHERE eg.grupo_id = %s
                  AND eg.fecha_fin IS NULL
                ORDER BY e.apellidos, e.nombre, e.id;
                """,
                [grupo_id],
            )
        else:
            if not (sede_id and grado_id):
                return []
            cur.execute(
                """
                SELECT DISTINCT e.id, e.apellidos, e.nombre, e.documento
                FROM public.estudiante_grupo eg
                JOIN public.grupos g   ON g.id = eg.grupo_id
                JOIN public.estudiantes e ON e.id = eg.estudiante_id
                WHERE g.sede_id  = %s
                  AND g.grado_id = %s
                  AND eg.fecha_fin IS NULL
                ORDER BY e.apellidos, e.nombre, e.id;
                """,
                [sede_id, grado_id],
            )

        for r in cur.fetchall():
            res.append({
                "id": r[0],
                "apellidos": r[1] or "",
                "nombre": r[2] or "",
                "documento": r[3] or "",
            })

    return res



def docente_boletin_grupo_pdf_alias(request):
    params = request.GET.copy()
    params["tipo"] = "grupo"
    params["formato"] = "pdf"
    return redirect(f"{reverse('rector_reportes_academicos_export')}?{params.urlencode()}")

def docente_boletin_grupo_excel_alias(request):
    params = request.GET.copy()
    params["tipo"] = "grupo"
    params["formato"] = "excel"
    return redirect(f"{reverse('rector_reportes_academicos_export')}?{params.urlencode()}")

def docente_boletin_estudiante_pdf_alias(request):
    params = request.GET.copy()
    params["tipo"] = "estudiante"
    params["formato"] = "pdf"
    return redirect(f"{reverse('rector_reportes_academicos_export')}?{params.urlencode()}")

def docente_boletin_estudiante_excel_alias(request):
    params = request.GET.copy()
    params["tipo"] = "estudiante"
    params["formato"] = "excel"
    return redirect(f"{reverse('rector_reportes_academicos_export')}?{params.urlencode()}")


# ==== Adaptador: build_boletines -> usa _dataset_boletines / _cargar_meta_grupo ====
def build_boletines(
    grupo_id: str | None,
    grado_id: str | None,
    sede_id: str | None,
    periodo_id: str | None,
    estudiante_id: str | None = None,
):
    grupo_id = (grupo_id or "").strip()
    periodo_id = (periodo_id or "").strip()
    estudiante_id = (estudiante_id or "").strip() or None
    if not grupo_id or not periodo_id:
        return []

    # ¿Qué periodo se está exportando?
    p_sel = {"1": "p1", "2": "p2", "3": "p3"}.get(str(periodo_id), "p3")

    estudiantes, _areas, _ = _dataset_boletines(grupo_id, periodo_id, estudiante_id)
    if not estudiantes:
        return []

    meta = _cargar_meta_grupo(grupo_id, periodo_id)

    boletines = []
    for s in estudiantes:
        rubrica = []
        for fila in s["detalle"]:
            # Normaliza tipo a lo que espera el template
            tipo = "area" if str(fila.get("tipo", "")).upper() == "AREA" else "asig"

            nombre = fila.get("nombre", "")

            # Notas crudas (en tu dataset pueden venir "" o número)
            p1 = fila.get("p1", "")
            p2 = fila.get("p2", "")
            p3 = fila.get("p3", "")
            fn = fila.get("final", "")
            fallas = fila.get("fallas", "")

            # Convierte "" -> None para cálculos
            n1 = None if p1 == "" else p1
            n2 = None if p2 == "" else p2
            n3 = None if p3 == "" else p3
            nf = None if fn == "" else fn

            # Nota del periodo visible (para "nivel" y "nivel_perdida")
            nota_periodo = {"p1": n1, "p2": n2, "p3": n3}[p_sel]
            nivel = _nivel_desempeno_rango(nota_periodo)

            # Flags de pérdida (regla general + especial de Inglés)
            p1_perdida = _es_perdida(nombre, n1) if n1 is not None else False
            p2_perdida = _es_perdida(nombre, n2) if n2 is not None else False
            p3_perdida = _es_perdida(nombre, n3) if n3 is not None else False
            final_perdida = _es_perdida(nombre, nf) if nf is not None else False
            nivel_perdida = _es_perdida(nombre, nota_periodo) if nota_periodo is not None else False

            rubrica.append({
                "tipo": tipo,                # "area" | "asig" (lo que el HTML compara)
                "nombre": nombre,
                "p1": p1,
                "p2": p2,
                "p3": p3,
                "final": fn,
                "fallas": fallas if tipo == "asig" else "",  # las áreas no muestran fallas

                # Nivel según el periodo seleccionado
                "nivel": nivel,

                # Flags que tu HTML usa para pintar rojo
                "p1_perdida": p1_perdida,
                "p2_perdida": p2_perdida,
                "p3_perdida": p3_perdida,
                "final_perdida": final_perdida,
                "nivel_perdida": nivel_perdida,

                # Extra opcional (mantienes compatibilidad)
                "logro_perdida": False,
                "reprobada": False,
            })

        boletines.append({
            "apellidos": s["apellidos"],
            "nombres": s["nombres"],
            "identificacion": s.get("documento", ""),
            "sede": meta["sede"],
            "jornada": meta.get("jornada", ""),
            "anio": meta.get("anio", ""),
            "periodo": meta.get("periodo_nombre", ""),
            "grado": meta["grado"],
            "grupo": meta["grupo"],
            "fecha": meta.get("fecha_emision", ""),

            "promedio": s.get("promedio", 0.0),
            "puesto_grupo": s.get("puesto_grupo", ""),
            "puesto_grado": s.get("puesto_grado", ""),
            "puesto_institucional": s.get("puesto_institucional", ""),
            "diagnostico": "",

            "rubrica": rubrica,
        })

    return boletines

# ========= Helpers de rol ADMINISTRATIVO =========
def _solo_admin(request) -> bool:
    """True si NO es rector ni docente (es decir, rol administrativo u otro que redirige a dashboard_admin)."""
    rol = (getattr(request.user, "rol", "") or "").upper()
    return rol not in ("RECTOR", "DOCENTE")

def _guard_admin(request: HttpRequest) -> HttpResponse | None:
    """Devuelve redirección si NO es administrativo; si es admin devuelve None."""
    if not _solo_admin(request):
        return _redir_por_rol(request.user)
    return None

# =========================================================
# APIs para filtros del rol ADMINISTRATIVO
# =========================================================
@login_required(login_url="login")
@require_GET
def api_admin_sedes(request: HttpRequest) -> JsonResponse:
    if (resp := _guard_admin(request)) is not None:
        return JsonResponse({"sedes": []})
    with connection.cursor() as cur:
        cur.execute("SELECT id, nombre FROM public.sedes ORDER BY nombre;")
        sedes = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"sedes": sedes})

@login_required(login_url="login")
@require_GET
def api_admin_grados_por_sede(request: HttpRequest) -> JsonResponse:
    if (resp := _guard_admin(request)) is not None:
        return JsonResponse({"grados": []})
    sede_id = (request.GET.get("sede_id") or "").strip()
    if not re.fullmatch(r"\d{1,10}", sede_id or ""):
        return JsonResponse({"grados": []})
    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT gr.id, gr.nombre
              FROM public.grupos g
              JOIN public.grados gr ON gr.id = g.grado_id
             WHERE g.sede_id = %s
             ORDER BY gr.id;
        """, [sede_id])
        grados = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"grados": grados})

@login_required(login_url="login")
@require_GET
def api_admin_grupos_por_sede_grado(request: HttpRequest) -> JsonResponse:
    if (resp := _guard_admin(request)) is not None:
        return JsonResponse({"grupos": []})
    sede_id  = (request.GET.get("sede_id")  or "").strip()
    grado_id = (request.GET.get("grado_id") or "").strip()
    ok = all(re.fullmatch(r"\d{1,10}", v or "") for v in (sede_id, grado_id))
    if not ok:
        return JsonResponse({"grupos": []})
    with connection.cursor() as cur:
        cur.execute("""
            SELECT id, nombre
              FROM public.grupos
             WHERE sede_id=%s AND grado_id=%s
             ORDER BY nombre;
        """, [sede_id, grado_id])
        grupos = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"grupos": grupos})

@login_required(login_url="login")
@require_GET
def api_admin_periodos_abiertos(request: HttpRequest) -> JsonResponse:
    if (resp := _guard_admin(request)) is not None:
        return JsonResponse({"periodos": []})
    with connection.cursor() as cur:
        cur.execute("""
            SELECT id, nombre
              FROM public.periodos
             WHERE abierto = TRUE
             ORDER BY fecha_inicio;
        """)
        periodos = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return JsonResponse({"periodos": periodos})

@login_required(login_url="login")
@require_GET
def api_admin_estudiantes_por_grupo(request: HttpRequest) -> JsonResponse:
    if (resp := _guard_admin(request)) is not None:
        return JsonResponse({"estudiantes": []})
    grupo_id = (request.GET.get("grupo_id") or "").strip()
    if not re.fullmatch(r"\d{1,10}", grupo_id or ""):
        return JsonResponse({"estudiantes": []})
    with connection.cursor() as cur:
        cur.execute("""
            SELECT e.id, e.nombre, e.apellidos, e.documento
              FROM public.estudiante_grupo eg
              JOIN public.estudiantes e ON e.id = eg.estudiante_id
             WHERE eg.grupo_id=%s AND eg.fecha_fin IS NULL
             ORDER BY e.apellidos, e.nombre;
        """, [grupo_id])
        estudiantes = [{"id": r[0], "nombre": r[1], "apellidos": r[2], "documento": r[3]} for r in cur.fetchall()]
    return JsonResponse({"estudiantes": estudiantes})

# Vistas HTML
@login_required
def rector_graficas_reportes(request):
    return render(request, "core/rector/graficas_reportes.html")

@login_required
def administrativo_graficas_reportes(request):
    return render(request, "core/administrativo/graficas_reportes.html")

# Helper seguro para leer ints
def _get_int(request, key):
    v = request.GET.get(key)
    try:
        return int(v) if v not in (None, "", "null") else None
    except ValueError:
        return None
    
# API: Estudiantes activos por sede (con filtros opcionales)
@login_required
def api_metrics_activos(request):
    """
    Activos por sede: cuenta estudiantes con matrícula activa (fecha_fin IS NULL)
    agrupando por sede_id del grupo. Filtros: sede (texto), grupo_id (int).
    """
    sede_nombre = request.GET.get("sede")
    grupo_id    = _get_int(request, "grupo_id")

    qs = EstudianteGrupo.objects.filter(fecha_fin__isnull=True)

    if sede_nombre:
        sede_ids = list(Sede.objects.filter(nombre=sede_nombre)
                                  .values_list("id", flat=True))
        if not sede_ids:
            return JsonResponse({"series": []})
        qs = qs.filter(grupo__sede_id__in=sede_ids)

    if grupo_id:
        qs = qs.filter(grupo_id=grupo_id)

    agregados = (
        qs.values("grupo__sede_id")
          .annotate(activos=Count("estudiante", distinct=True))
    )

    nombres_sede = {s.id: s.nombre for s in Sede.objects.all()}
    data = [{"name": nombres_sede.get(r["grupo__sede_id"], f"Sede {r['grupo__sede_id']}"),
             "value": r["activos"] or 0}
            for r in agregados]
    data.sort(key=lambda x: x["name"])

    return JsonResponse({"series": data})

def _to_int_or_none(v):
    try:
        return int(v) if v not in (None, "", "null") else None
    except (TypeError, ValueError):
        return None

@login_required
def api_metrics_reprobados(request):
    """
    Devuelve conteos de REPROBADOS (< umbral) por ASIGNATURA dentro de un grupo.
    Filtros soportados (todos opcionales):
      - sede (nombre de la sede)
      - grado_id
      - grupo_id
      - periodo_id
      - threshold (float, default 3.0)
    Respuesta: { "series": [ {"name": "Matemáticas", "value": 12}, ... ] }
    """
    sede_nombre = request.GET.get("sede")              # string
    grado_id    = _to_int_or_none(request.GET.get("grado_id"))
    grupo_id    = _to_int_or_none(request.GET.get("grupo_id"))
    periodo_id  = _to_int_or_none(request.GET.get("periodo_id"))

    # Umbral de aprobación (por defecto 3.0). Acepta coma o punto.
    th_raw = request.GET.get("threshold")
    try:
        threshold = float((th_raw or "3.0").replace(",", "."))
    except ValueError:
        threshold = 3.0

    where = ["eg.fecha_fin IS NULL"]  # solo alumnos activos en su grupo
    params = []

    if sede_nombre:
        where.append("s.nombre = %s")
        params.append(sede_nombre)

    if grado_id:
        where.append("g.grado_id = %s")
        params.append(grado_id)

    if grupo_id:
        where.append("g.id = %s")
        params.append(grupo_id)

    if periodo_id:
        where.append("n.periodo_id = %s")
        params.append(periodo_id)

    # Reprobados por asignatura
    where.append("n.nota < %s")
    params.append(threshold)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    sql = f"""
        SELECT a.id, a.nombre AS asignatura, COUNT(*) AS reprobados
        FROM notas n
        JOIN estudiante_grupo eg ON eg.estudiante_id = n.estudiante_id
        JOIN grupos g           ON g.id = eg.grupo_id
        LEFT JOIN sedes s       ON s.id = g.sede_id
        JOIN asignaturas a      ON a.id = n.asignatura_id
        {where_sql}
        GROUP BY a.id, a.nombre
        ORDER BY a.nombre;
    """

    with connection.cursor() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()

    data = [{"name": r[1], "value": int(r[2] or 0)} for r in rows]
    return JsonResponse({"series": data, "threshold": threshold})

# ==========================
# MÉTRICA: HISTOGRAMA DE NOTAS
# ==========================
@login_required
@require_GET
def api_metrics_histograma(request):
    """
    Devuelve un histograma de notas en “bins” de 0.5:
    1.0, 1.5, 2.0, ..., 5.0

    Filtros (opcionales):
      - sede       (por NOMBRE de la sede, como viene del frontend)
      - grupo_id   (id del grupo)
      - periodo_id (id del periodo)
    """
    sede_nombre = (request.GET.get("sede") or "").strip()
    grupo_id    = (request.GET.get("grupo_id") or "").strip()
    periodo_id  = (request.GET.get("periodo_id") or "").strip()

    # Construir WHERE dinámico
    wheres = []
    params = []

    # Solo notas vigentes del estudiante en el grupo (eg.fecha_fin IS NULL)
    base_sql = """
        SELECT
            ROUND(n.nota * 2) / 2.0 AS bucket,  -- bins de 0.5
            COUNT(*) AS cnt
        FROM public.notas n
        JOIN public.estudiante_grupo eg
              ON eg.estudiante_id = n.estudiante_id
             AND eg.fecha_fin IS NULL
        JOIN public.grupos g
              ON g.id = eg.grupo_id
        LEFT JOIN public.sedes s
              ON s.id = g.sede_id
    """

    if periodo_id:
        wheres.append("n.periodo_id = %s")
        params.append(periodo_id)

    if grupo_id:
        wheres.append("g.id = %s")
        params.append(grupo_id)

    if sede_nombre:
        # Filtra por NOMBRE de sede (coincide con lo que envía el front)
        wheres.append("s.nombre = %s")
        params.append(sede_nombre)

    if wheres:
        base_sql += " WHERE " + " AND ".join(wheres)

    base_sql += """
        GROUP BY bucket
        ORDER BY bucket
    """

    # Ejecutar y mapear resultados
    with connection.cursor() as cur:
        cur.execute(base_sql, params)
        rows = cur.fetchall()  # [(bucket, cnt), ...]

    # Queremos bins fijos desde 1.0 hasta 5.0 (0.5 en 0.5)
    # Si tu escala arranca en 0.0, ajusta el rango.
    all_bins = [x / 2.0 for x in range(2, 11)]  # 1.0..5.0 (2->10)/2
    counts = {float(b): 0 for b in all_bins}
    for b, c in rows:
        # b llega como Decimal/float; normalizamos a float con 1 decimal
        try:
            bkey = float(b)
        except:
            continue
        if bkey in counts:
            counts[bkey] = int(c)

    series = [{"name": f"{b:.1f}", "value": counts[b]} for b in all_bins]

    return JsonResponse({"series": series})


def user_is_rector(user):
    return hasattr(user, "rol") and str(user.rol).upper() == "RECTOR"


def rector_required(view_func):
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated or not user_is_rector(request.user):
            return HttpResponseForbidden("Solo para usuarios con rol Rector.")
        return view_func(request, *args, **kwargs)
    return _wrapped


@never_cache
@login_required
@rector_required
@require_GET
def planillas_export_landing(request):
    formato = (request.GET.get("formato") or "").lower()
    sede = request.GET.get("sede") or ""
    grado = request.GET.get("grado") or ""
    grupo = request.GET.get("grupo") or ""

    base_qs = f"sede={sede}&grado={grado}&grupo={grupo}"
    url_pdf = f"{reverse('core:planillas_export_pdf')}?{base_qs}"
    url_excel = f"{reverse('core:planillas_export_excel')}?{base_qs}"

    contexto = {
        "url_pdf": url_pdf,
        "url_excel": url_excel,
        "sel_sede": sede or "Todas",
        "sel_grado": grado or "Todos",
        "sel_grupo": grupo or "Todos",
        "formato": formato,
    }
    return render(request, "core/rector/landing.html", contexto)


# =========================
#  LISTADO / FILTROS (UI)
# =========================
@login_required
@rector_required
def planillas_index(request):
    """
    Pantalla con filtros (sede, grado, grupo) y botones de descarga.
    """
    sede  = (request.GET.get("sede") or "").strip()
    grado = (request.GET.get("grado") or "").strip()
    grupo = (request.GET.get("grupo") or "").strip()

    with connection.cursor() as cur:
        cur.execute("SELECT id, nombre FROM public.sedes ORDER BY nombre;")
        sedes = cur.fetchall()

        cur.execute("SELECT id, nombre FROM public.grados ORDER BY nombre;")
        grados = cur.fetchall()

        cur.execute("""
            SELECT g.id, CONCAT(s.nombre, ' - ', gr.nombre, ' - ', g.nombre)
            FROM public.grupos g
            JOIN public.sedes  s  ON s.id  = g.sede_id
            JOIN public.grados gr ON gr.id = g.grado_id
            ORDER BY s.nombre, gr.nombre, g.nombre;
        """)
        grupos = cur.fetchall()

    return render(
        request,
        "core/rector/planillas.html",
        {
            "sedes": sedes, "grados": grados, "grupos": grupos,
            "sel_sede": sede, "sel_grado": grado, "sel_grupo": grupo,
        },
    )


# ======================================
#  LANDING DE EXPORTACIÓN (REDIRECCIÓN)
# ======================================
@login_required
@rector_required
@require_GET
def planillas_export_landing(request):
    """
    Recibe ?formato=pdf|excel + filtros y redirige a la URL concreta
    de exportación manteniendo los parámetros. **Sin namespace `core:`**.
    """
    formato = (request.GET.get("formato") or "pdf").lower()
    sede    = (request.GET.get("sede") or "").strip()
    grado   = (request.GET.get("grado") or "").strip()
    grupo   = (request.GET.get("grupo") or "").strip()

    base_qs = urlencode({"sede": sede, "grado": grado, "grupo": grupo})

    if formato == "excel":
        url = f"{reverse('planillas_export_excel')}?{base_qs}"
    else:
        url = f"{reverse('planillas_export_pdf')}?{base_qs}"

    return redirect(url)


# =========================
#  EXPORTACIÓN A EXCEL
# =========================
@login_required
@rector_required
@require_GET
def planillas_export_excel(request):
    sede  = (request.GET.get("sede") or "").strip()
    grado = (request.GET.get("grado") or "").strip()
    grupo = (request.GET.get("grupo") or "").strip()

    # Validación segura de params numéricos
    for v in (sede, grado, grupo):
        if v and not re.fullmatch(r"\d{1,10}", v):
            return HttpResponse("Parámetros inválidos.", status=400)

    # Consulta
    with connection.cursor() as cur:
        cur.execute("""
            SELECT e.apellidos, e.nombre, e.documento,
                   gr.nombre AS grado, g.nombre AS grupo
            FROM public.estudiante_grupo eg
            JOIN public.estudiantes e ON e.id = eg.estudiante_id
            JOIN public.grupos g ON g.id = eg.grupo_id
            JOIN public.grados gr ON gr.id = g.grado_id
            LEFT JOIN public.sedes s ON s.id = g.sede_id
            WHERE eg.fecha_fin IS NULL
              AND (%s = '' OR s.id::text = %s)
              AND (%s = '' OR gr.id::text = %s)
              AND (%s = '' OR g.id::text = %s)
            ORDER BY e.apellidos, e.nombre;
        """, [sede, sede, grado, grado, grupo, grupo])
        filas = cur.fetchall()

        cur.execute(
            "SELECT COALESCE((SELECT nombre FROM public.sedes WHERE id::text=%s), 'Todas');",
            [sede or ""],
        )
        header_sede = cur.fetchone()[0]

        cur.execute(
            "SELECT COALESCE((SELECT nombre FROM public.grados WHERE id::text=%s), 'Todos');",
            [grado or ""],
        )
        header_grado = cur.fetchone()[0]

        cur.execute(
            "SELECT COALESCE((SELECT nombre FROM public.grupos WHERE id::text=%s), 'Todos');",
            [grupo or ""],
        )
        header_grupo = cur.fetchone()[0]

    titulo = f"Planilla - Sede: {header_sede} | Grado: {header_grado} | Grupo: {header_grupo}"

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes activos"

    ws.merge_cells("A1:E1")
    cell_title = ws["A1"]
    cell_title.value = titulo
    cell_title.font = Font(size=14, bold=True)
    cell_title.alignment = Alignment(horizontal="center")

    ws.append([])  # fila en blanco
    headers = ["Apellidos", "Nombres", "Documento", "Grado", "Grupo"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=3, column=col_idx).font = Font(bold=True)

    for ap, no, doc, gr_nombre, g_nombre in filas:
        ws.append([ap, no, doc, gr_nombre, g_nombre])

    # Auto width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    now_str = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"planilla_estudiantes_{now_str}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# =========================
#  EXPORTACIÓN A PDF
# =========================
@login_required
@rector_required
@require_GET
def planillas_export_pdf(request):
    sede  = (request.GET.get("sede") or "").strip()
    grado = (request.GET.get("grado") or "").strip()
    grupo = (request.GET.get("grupo") or "").strip()

    for v in (sede, grado, grupo):
        if v and not re.fullmatch(r"\d{1,10}", v):
            return HttpResponse("Parámetros inválidos.", status=400)

    with connection.cursor() as cur:
        cur.execute("""
            SELECT e.apellidos, e.nombre, e.documento,
                   gr.nombre AS grado, g.nombre AS grupo
            FROM public.estudiante_grupo eg
            JOIN public.estudiantes e ON e.id = eg.estudiante_id
            JOIN public.grupos g ON g.id = eg.grupo_id
            JOIN public.grados gr ON gr.id = g.grado_id
            LEFT JOIN public.sedes s ON s.id = g.sede_id
            WHERE eg.fecha_fin IS NULL
              AND (%s = '' OR s.id::text = %s)
              AND (%s = '' OR gr.id::text = %s)
              AND (%s = '' OR g.id::text = %s)
            ORDER BY e.apellidos, e.nombre;
        """, [sede, sede, grado, grado, grupo, grupo])
        filas = cur.fetchall()

        cur.execute(
            "SELECT COALESCE((SELECT nombre FROM public.sedes WHERE id::text=%s), 'Todas');",
            [sede or ""],
        )
        header_sede = cur.fetchone()[0]
        cur.execute(
            "SELECT COALESCE((SELECT nombre FROM public.grados WHERE id::text=%s), 'Todos');",
            [grado or ""],
        )
        header_grado = cur.fetchone()[0]
        cur.execute(
            "SELECT COALESCE((SELECT nombre FROM public.grupos WHERE id::text=%s), 'Todos');",
            [grupo or ""],
        )
        header_grupo = cur.fetchone()[0]

    # Render HTML para PDF
    html = render_to_string(
        "core/rector/pdf.html",
        {
            "filas": filas,
            "header_sede": header_sede,
            "header_grado": header_grado,
            "header_grupo": header_grupo,
            "generado": timezone.now(),
        },
    )

    pdf_bytes = HTML(string=html).write_pdf(
        stylesheets=[CSS(string="""
            @page { size: A4; margin: 18mm; }
            h1 { font-size: 16pt; margin: 0 0 10px 0; }
            .meta { font-size: 10pt; margin-bottom: 8px; color: #555; }
            table { width:100%; border-collapse: collapse; font-size: 10pt; }
            th, td { border: 1px solid #ddd; padding: 6px 8px; }
            th { background: #f2f2f2; text-align: left; }
            tr:nth-child(even) td { background: #fafafa; }
        """)]
    )

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="planilla_estudiantes.pdf"'
    return response

@login_required
@rector_required
@require_GET
def api_grados_por_sede(request):
    """
    Devuelve los GRADOS que existen para la sede dada,
    detectándolos según los grupos creados en esa sede.
    """
    sede_id = (request.GET.get("sede_id") or "").strip()

    with connection.cursor() as cur:
        if sede_id:
            cur.execute("""
                SELECT DISTINCT gr.id, gr.nombre
                FROM public.grupos g
                JOIN public.grados gr ON gr.id = g.grado_id
                WHERE g.sede_id::text = %s
                ORDER BY gr.nombre;
            """, [sede_id])
        else:
            # Si no hay sede, puedes devolver todos o ninguno.
            # Aquí devolvemos todos (útil para pre-cargar).
            cur.execute("SELECT id, nombre FROM public.grados ORDER BY nombre;")
        rows = cur.fetchall()

    data = [{"id": str(i), "nombre": n} for (i, n) in rows]
    return JsonResponse({"results": data})


@login_required
@rector_required
@require_GET
def api_grupos_por_sede_grado(request):
    """
    Devuelve los GRUPOS de una sede (opcionalmente filtrados por grado).
    """
    sede_id = (request.GET.get("sede_id") or "").strip()
    grado_id = (request.GET.get("grado_id") or "").strip()

    with connection.cursor() as cur:
        cur.execute("""
            SELECT g.id,
                   CONCAT(s.nombre, ' - ', gr.nombre, ' - ', g.nombre) AS full
            FROM public.grupos g
            JOIN public.sedes s  ON s.id  = g.sede_id
            JOIN public.grados gr ON gr.id = g.grado_id
            WHERE (%s = '' OR g.sede_id::text = %s)
              AND (%s = '' OR g.grado_id::text = %s)
            ORDER BY s.nombre, gr.nombre, g.nombre;
        """, [sede_id, sede_id, grado_id, grado_id])
        rows = cur.fetchall()

    data = [{"id": str(i), "full": f} for (i, f) in rows]
    return JsonResponse({"results": data})

@rector_required
@require_http_methods(["GET", "POST"])
def rector_eliminar_estudiante(request):
    """
    Página para que el Rector elimine estudiantes por documento.
    - GET: muestra el formulario.
    - POST: elimina (con respaldo previo en estudiantes_borrados).
    """
    if request.method == "POST":
        documento = (request.POST.get("documento") or "").strip()
        if not re.fullmatch(r"\d{5,20}", documento):
            messages.error(request, "Documento inválido. Debe tener solo dígitos (5–20).")
            return redirect("rector_eliminar_estudiante")

        try:
            with transaction.atomic():
                with connection.cursor() as cur:
                    # 1) Obtener estudiante por documento
                    cur.execute("""
                        SELECT e.id, e.nombre, e.apellidos, e.sede_id, COALESCE(s.nombre, '')
                        FROM public.estudiantes e
                        LEFT JOIN public.sedes s ON s.id = e.sede_id
                        WHERE e.documento = %s
                        LIMIT 1
                    """, [documento])
                    row = cur.fetchone()

                    if not row:
                        messages.error(request, f"No existe estudiante con documento {documento}.")
                        return redirect("rector_eliminar_estudiante")

                    est_id, est_nombre, est_apellidos, est_sede_id, est_sede_nombre = row

                    # 2) Obtener snapshot de su último grupo activo (si existe)
                    cur.execute("""
                        SELECT s.nombre AS sede, gr.nombre AS grado, g.nombre AS grupo
                        FROM public.estudiante_grupo eg
                        JOIN public.grupos g  ON g.id  = eg.grupo_id
                        JOIN public.grados gr ON gr.id = g.grado_id
                        JOIN public.sedes  s  ON s.id  = g.sede_id
                        WHERE eg.estudiante_id = %s
                          AND eg.fecha_fin IS NULL
                        ORDER BY eg.id DESC
                        LIMIT 1
                    """, [est_id])
                    last_grp = cur.fetchone()

                    sede_snap  = (last_grp[0] if last_grp else est_sede_nombre) or None
                    grado_snap = (last_grp[1] if last_grp else None)
                    grupo_snap = (last_grp[2] if last_grp else None)

                    # 3) Guardar respaldo en estudiantes_borrados (auditoría)
                    cur.execute("""
                        INSERT INTO public.estudiantes_borrados
                          (estudiante_id, documento, nombre, apellidos, sede_id, sede,
                           grado, grupo, eliminado_por, eliminado_en, extra)
                        VALUES
                          (%s, %s, %s, %s, %s, %s,
                           %s, %s, %s, now(),
                           jsonb_build_object('ip', %s, 'ua', %s))
                    """, [
                        est_id, documento, est_nombre, est_apellidos,
                        est_sede_id, sede_snap, grado_snap, grupo_snap,
                        request.user.username,
                        request.META.get("REMOTE_ADDR"),
                        request.META.get("HTTP_USER_AGENT"),
                    ])

                    # 4) Borrar datos dependientes y estudiante
                    #    (historial primero; notas luego; estudiante al final.
                    #     estudiante_grupo cae por CASCADE según tu esquema)
                    cur.execute("DELETE FROM public.notas_historial WHERE estudiante_id=%s;", [est_id])
                    cur.execute("DELETE FROM public.notas           WHERE estudiante_id=%s;", [est_id])
                    cur.execute("DELETE FROM public.estudiantes     WHERE id=%s;", [est_id])
                    est_borrados = cur.rowcount or 0

            if est_borrados:
                messages.success(
                    request,
                    f"Estudiante {est_apellidos} {est_nombre} (doc: {documento}) eliminado correctamente."
                )
            else:
                messages.warning(request, "No se pudo eliminar el estudiante (ya no existía).")

        except Exception as e:
            messages.error(request, f"No se pudo eliminar: {e}")

        return redirect("rector_eliminar_estudiante")

    # GET
    return render(request, "core/rector/eliminar_estudiante.html")


@login_required
@rector_required
@require_GET
def api_estudiante_por_documento(request):
    """
    Devuelve datos del estudiante y su grupo ACTUAL (si tiene) por documento.
    Respuesta JSON para el preview en la vista.
    """
    documento = (request.GET.get("doc") or "").strip()
    if not documento:
        return JsonResponse({"ok": False, "error": "Falta parámetro doc"}, status=400)

    if not re.fullmatch(r"\d{5,20}", documento):
        return JsonResponse({"ok": False, "error": "Documento inválido"}, status=400)

    with connection.cursor() as cur:
        cur.execute("""
            SELECT
                e.id,
                e.documento,
                e.nombre,
                e.apellidos,
                s.nombre AS sede,
                gr.nombre AS grado,
                g.nombre  AS grupo
            FROM public.estudiantes e
            LEFT JOIN public.estudiante_grupo eg
                   ON eg.estudiante_id = e.id
                  AND eg.fecha_fin IS NULL
            LEFT JOIN public.grupos g ON g.id = eg.grupo_id
            LEFT JOIN public.grados gr ON gr.id = g.grado_id
            LEFT JOIN public.sedes  s  ON s.id  = g.sede_id
            WHERE e.documento = %s
            LIMIT 1;
        """, [documento])
        row = cur.fetchone()

    if not row:
        return JsonResponse({"ok": False, "found": False})

    est_id, doc, nombre, apellidos, sede, grado, grupo = row
    return JsonResponse({
        "ok": True,
        "found": True,
        "estudiante": {
            "id": est_id,
            "documento": doc,
            "nombre": nombre,
            "apellidos": apellidos,
            "sede": sede or "-",
            "grado": grado or "-",
            "grupo": grupo or "-",
        }
    })